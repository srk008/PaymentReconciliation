from flask import Flask, render_template, request, redirect, url_for, send_from_directory, flash, jsonify, json
import mysql.connector
from mysql.connector import Error
import config
import openpyxl
import os
import re
import ast
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font

# this is server
app = Flask(__name__)
app.secret_key = '9f2a1b8c4d6e7f0123456789abcdef01'

str_pattern_reg = "^[A-Za-z ]+$"

class ResponseHandler:
    @staticmethod
    def invalid_name(entity):
        return {'status': 'error', 'message': f'Invalid {entity} name. Only letters are allowed!'}

    @staticmethod
    def already_exists(entity):
        return {'status': 'exists', 'message': f'{entity.capitalize()} already exists!'}

    @staticmethod
    def add_success(entity):
        return {'status': 'success', 'message': f'{entity.capitalize()} added successfully!'}

    @staticmethod
    def add_failure(entity):
        return {'status': 'error', 'message': f'Failed to add {entity}.'}

    @staticmethod
    def is_available(entity):
        return {'status': 'available', 'message': f'{entity.capitalize()} name is available!'}

    @staticmethod
    def delete_success(entity):
        return {'status': 'success', 'message': f'{entity.capitalize()} deleted successfully!'}

    @staticmethod
    def delete_failure(entity):
        return {'status': 'error', 'message': f'Failed to delete {entity}.'}

    @staticmethod
    def update_success(entity):
        return {'status': 'success', 'message': f'{entity.capitalize()} updated successfully!'}

    @staticmethod
    def update_failure(entity):
        return {'status': 'error', 'message': f'Failed to update {entity}.'}

    @staticmethod
    def fetch_failure(entity):
        return f"Failed to fetch {entity}"


# Helper: JSON Response Formatter
def json_response(message_obj, status_code):
    return jsonify(message_obj), status_code


# this is Index page OR Home page..
@app.route('/')
def index():
    return render_template('index.html')


# this is Profile page ..
@app.route('/admin_profile', methods=['GET', 'POST'])
def admin_profile():
    return render_template('admin_profile.html')


# ------------------------- State controller ------------------------------------------
@app.route('/add_state', methods=['GET', 'POST'])
def add_state():
    connection = config.get_db_connection()
    statedata = []

    if connection:
        cursor = connection.cursor()
        if request.method == 'POST':
            state_name = request.form['state_Name'].strip()

            if not re.match(str_pattern_reg, state_name):
                return json_response(ResponseHandler.invalid_name("state"), 400)

            try:
                cursor.execute("SELECT * FROM states WHERE State_Name = %s", (state_name,))
                if cursor.fetchone():
                    return json_response(ResponseHandler.already_exists("state"), 409)

                # cursor.callproc("CheckStateExists", (state_name,))
                # for data in cursor.stored_results():
                #     existing_state = data.fetchone()
                #
                # if existing_state:
                #     return json_response(ResponseHandler.already_exists("state"), 409)

                # cursor.execute("call SaveState (%s)", (state_name,))
                cursor.callproc("SaveState", (state_name,))
                connection.commit()
                return json_response(ResponseHandler.add_success("state"), 200)

            except mysql.connector.Error as e:
                print(f"Error inserting state: {e}")
                return json_response(ResponseHandler.add_failure("state"), 500)

        try:
            # cursor.execute("SELECT State_ID, State_Name FROM states")
            # statedata = cursor.fetchall()
            cursor.callproc("GetAllStates")
            for res in cursor.stored_results():
                statedata = res.fetchall()
        except mysql.connector.Error as e:
            print(f"Error fetching states: {e}")
            return ResponseHandler.fetch_failure("states"), 500
        finally:
            cursor.close()
            connection.close()

    return render_template('add_state.html', statedata=statedata)


# AJAX route to check state existence
@app.route('/check_state', methods=['POST'])
def check_state():
    connection = config.get_db_connection()


    if connection:
        cursor = connection.cursor()
        state_name = request.json.get('state_Name', '').strip()

        if not re.match(str_pattern_reg, state_name):
            return json_response(ResponseHandler.invalid_name("state"), 400)

        try:
            cursor.execute("SELECT * FROM states WHERE State_Name = %s", (state_name,))
            existing_state = cursor.fetchone()

            # cursor.callproc("CheckStateExists", (state_name,))
            # for data in cursor.stored_results():
            #     existing_state = data.fetchone()

            if existing_state:
                return json_response(ResponseHandler.already_exists("state"), 409)
            else:
                return json_response(ResponseHandler.is_available("state"), 200)

        except mysql.connector.Error as e:
            print(f"Error checking state: {e}")
            return json_response(ResponseHandler.add_failure("state"), 500)
        finally:
            cursor.close()
            connection.close()


# Delete State
@app.route('/delete_state/<int:id>', methods=['GET'])
def deleteState(id):
    connection = config.get_db_connection()
    cursor = connection.cursor()

    try:
        # cursor.execute("DELETE FROM states WHERE State_ID = %s", (id,))
        cursor.callproc('DeleteState',(id,))
        connection.commit()
        # For API response
        # return json_response(ResponseHandler.delete_success("state"), 200)

    except mysql.connector.Error as e:
        print(f"Error deleting data: {e}")
        return json_response(ResponseHandler.delete_failure("state"), 500)

    finally:
        cursor.close()
        connection.close()

    return redirect(url_for('add_state'))


# Edit State
@app.route('/edit_state/<int:id>', methods=['GET', 'POST'])
def editState(id):
    connection = config.get_db_connection()
    cursor = connection.cursor()
    # str_pattern_reg = r"^[A-Za-z\s]+$"

    if request.method == 'POST':
        state_name = request.form['state_Name'].strip()

        if not re.match(str_pattern_reg, state_name):
            return ResponseHandler.invalid_name("state"), 400

        try:
            # cursor.execute("UPDATE states SET State_Name = %s WHERE State_ID = %s", (state_name, id))
            cursor.callproc("UpdateStateById",(id,state_name))
            connection.commit()
            return redirect(url_for('add_state'))
        except mysql.connector.Error as e:
            print(f"Error updating data: {e}")
            return ResponseHandler.add_failure("state"), 500
        finally:
            cursor.close()
            connection.close()

    try:
        cursor.execute("SELECT * FROM states WHERE State_ID = %s", (id,))
        state = cursor.fetchone()
        if state is None:
            return "State not found", 404
    except mysql.connector.Error as e:
        print(f"Error retrieving data: {e}")
        return ResponseHandler.fetch_failure("state"), 500
    finally:
        cursor.close()
        connection.close()

    return render_template('edit_state.html', state=state)


# -------- end State controller -----------

# ------------------------- District controller ------------------------------------------
@app.route('/add_district', methods=['GET', 'POST'])
def add_district():
    connection = config.get_db_connection()
    districtdata = []
    states = []

    if connection:
        cursor = connection.cursor()

        try:
            # cursor.execute("SELECT State_ID, State_Name FROM states")
            # states = cursor.fetchall()
            cursor.callproc("GetAllStates")
            for res in cursor.stored_results():
                states = res.fetchall()

        except mysql.connector.Error as e:
            print(f"Error fetching states: {e}")
            return ResponseHandler.fetch_failure("states"), 500

        if request.method == 'POST':
            district_name = request.form['district_Name'].strip()
            state_id = request.form['state_Id']

            if not re.match(str_pattern_reg, district_name):
                return json_response(ResponseHandler.invalid_name("district"), 400)

            try:
                cursor.execute("SELECT * FROM districts WHERE District_Name = %s AND State_Id = %s",
                               (district_name, state_id))
                if cursor.fetchone():
                    return json_response(ResponseHandler.already_exists("district"), 409)

                cursor.callproc('SaveDistrict', (district_name, state_id))
                connection.commit()

                return json_response(ResponseHandler.add_success("district"), 200)

            except mysql.connector.Error as e:
                print(f"Error inserting district: {e}")
                return json_response(ResponseHandler.add_failure("district"), 500)

        try:
            # cursor.execute("SELECT d.District_id, d.District_Name, s.State_Name, s.State_Id FROM districts d JOIN states s ON d.State_Id = s.State_ID")
            # districtdata = cursor.fetchall()
            cursor.callproc("GetAllDistricts")
            for dis in cursor.stored_results():
                districtdata = dis.fetchall()
        except mysql.connector.Error as e:
            print(f"Error fetching districts: {e}")
            return ResponseHandler.fetch_failure("districts"), 500
        finally:
            cursor.close()
            connection.close()

    return render_template('add_district.html', districtdata=districtdata, states=states)


# AJAX route to check district existence
@app.route('/check_district', methods=['POST'])
def check_district():
    connection = config.get_db_connection()

    if connection:
        cursor = connection.cursor()
        district_name = request.json.get('district_Name', '').strip()
        state_id = request.json.get('state_Id', '')

        if not re.match(str_pattern_reg, district_name):
            return json_response(ResponseHandler.invalid_name("district"), 400)

        try:
            cursor.execute("SELECT * FROM districts WHERE District_Name = %s AND State_Id = %s",
                           (district_name, state_id))
            existing_district = cursor.fetchone()

            if existing_district:
                return json_response(ResponseHandler.already_exists("district"), 409)
            else:
                return json_response(ResponseHandler.is_available("district"), 200)

        except mysql.connector.Error as e:
            print(f"Error checking district: {e}")
            return json_response(ResponseHandler.add_failure("district"), 500)
        finally:
            cursor.close()
            connection.close()


# this is delete District method by id..
@app.route('/delete_district/<int:district_id>', methods=['GET', 'POST'])
def delete_district(district_id):
    connection = config.get_db_connection()

    if connection:
        cursor = connection.cursor()
        try:
            # cursor.execute("DELETE FROM districts WHERE District_id = %s", (district_id,))
            cursor.callproc("DeleteDistrict",(district_id,))
            connection.commit()
        except mysql.connector.Error as e:
            print(f"Error deleting district: {e}")
            return json_response(ResponseHandler.delete_failure("district"), 500)
        finally:
            cursor.close()
            connection.close()

    return redirect('/add_district')


# this is update District page by id ..
@app.route('/edit_district/<int:district_id>', methods=['GET', 'POST'])
def edit_district(district_id):
    connection = config.get_db_connection()
    districtdata = []
    states = []

    if connection:
        cursor = connection.cursor()

        # Retrieve all states for dropdown
        try:
            # cursor.execute("SELECT State_ID, State_Name FROM states")
            # states = cursor.fetchall()
            cursor.callproc("GetAllStates")
            for res in cursor.stored_results():
                states = res.fetchall()

        except mysql.connector.Error as e:
            print(f"Error fetching states: {e}")
            return ResponseHandler.fetch_failure("states"), 500

        # Retrieve district info
        try:
            cursor.execute("SELECT District_Name, State_Id FROM districts WHERE District_id = %s", (district_id,))
            districtdata = cursor.fetchone()
        except mysql.connector.Error as e:
            print(f"Error fetching district data: {e}")
            return ResponseHandler.fetch_failure("district"), 500

        # Handle update
        if request.method == 'POST':
            district_name = request.form['district_Name']
            state_id = request.form['state_Id']

            try:
                # cursor.execute( "UPDATE districts SET District_Name = %s, State_Id = %s WHERE District_id = %s",
                #     (district_name, state_id, district_id) )

                cursor.callproc("UpdateDistrict", (district_id, state_id, district_name,))
                connection.commit()
            except mysql.connector.Error as e:
                print(f"Error updating district: {e}")
                return ResponseHandler.update_failure("district"), 500
            return redirect('/add_district')

    return render_template('edit_district.html', districtdata=districtdata, states=states)
# --------- end District controller -------------

# ------------------------- Block controller ------------------------------------------
@app.route('/add_block', methods=['GET', 'POST'])
def add_block():
    connection = config.get_db_connection()
    block_data = []
    states = []

    if connection:
        cursor = connection.cursor()
        try:
            # cursor.execute("SELECT State_ID, State_Name FROM states")
            # states = cursor.fetchall()
            cursor.callproc("GetAllStates")
            for res in cursor.stored_results():
                states = res.fetchall()

        except mysql.connector.Error as e:
            print(f"Error fetching states: {e}")
            return json_response(ResponseHandler.fetch_failure("states"), 500)

        if request.method == 'POST':
            block_name = request.form['block_Name'].strip()
            district_id = request.form['district_Id']

            if not re.match(str_pattern_reg, block_name):
                return json_response(ResponseHandler.invalid_name("block"), 400)

            try:
                cursor.execute("SELECT * FROM blocks WHERE Block_Name = %s AND District_id = %s",
                               (block_name, district_id))
                existing_block = cursor.fetchone()

                if existing_block:
                    return json_response(ResponseHandler.already_exists("block"), 409)

                cursor.callproc('SaveBlock', (block_name, district_id))
                connection.commit()

                return json_response(ResponseHandler.add_success("block"), 200)

            except mysql.connector.Error as e:
                print(f"Error adding block: {e}")
                return json_response(ResponseHandler.add_failure("block"), 500)

        # Fetch all blocks to display
        try:
            # cursor.execute(
            #     """SELECT b.Block_Id, b.Block_Name, d.District_Name
            #        FROM blocks b
            #        JOIN districts d ON b.District_id = d.District_id"""
            # )
            # block_data = cursor.fetchall()
            cursor.callproc("GetAllBlocks")
            for blocks in cursor.stored_results():
                block_data = blocks.fetchall()

        except mysql.connector.Error as e:
            print(f"Error fetching blocks: {e}")
            return json_response(ResponseHandler.fetch_failure("blocks"), 500)
        finally:
            cursor.close()
            connection.close()
    return render_template('add_block.html', block_data=block_data, states=states)


# check block
@app.route('/check_block', methods=['POST'])
def check_block():
    connection = config.get_db_connection()
    cursor = connection.cursor()
    block_name = request.json.get('block_Name', '').strip()
    district_id = request.json.get('district_Id', '')

    if not re.match(str_pattern_reg, block_name):
        return json_response(ResponseHandler.invalid_name("block"), 400)

    cursor.execute("SELECT * FROM blocks WHERE Block_Name = %s AND District_id = %s", (block_name, district_id))
    existing_block = cursor.fetchone()

    if existing_block:
        return json_response(ResponseHandler.already_exists("block"), 409)

    return json_response(ResponseHandler.is_available("block"), 200)


# update block by id
# @app.route('/edit_block/<int:block_id>', methods=['GET', 'POST'])
# def edit_block(block_id):
#     connection = config.get_db_connection()
#     block_data, states, districts = [], [], []
#
#     if connection:
#         cursor = connection.cursor()
#         try:
#             # cursor.execute("SELECT State_ID, State_Name FROM states")
#             # states = cursor.fetchall()
#             cursor.callproc("GetAllStates")
#             for res in cursor.stored_results():
#                 states = res.fetchall()
#
#             # cursor.execute("SELECT Block_Name, District_id FROM blocks WHERE Block_Id = %s", (block_id,))
#             # block_data = cursor.fetchone()
#             cursor.callproc("GetBlockById",(block_id,))
#             for block in cursor.stored_results():
#                 block_data = block.fetchone()
#
#         except mysql.connector.Error as e:
#             print(f"Error fetching block data: {e}")
#             return json_response(ResponseHandler.fetch_failure("block data"), 500)
#
#         if request.method == 'POST':
#             block_name = request.form['block_Name']
#             district_id = request.form['district_Id']
#
#             try:
#                 # cursor.execute("UPDATE blocks SET Block_Name = %s, District_id = %s WHERE Block_Id = %s",
#                 #                (block_name, district_id, block_id))
#
#                 cursor.callproc("UpdateBlock",(block_id,block_name, district_id))
#                 connection.commit()
#             except mysql.connector.Error as e:
#                 print(f"Error updating block: {e}")
#                 return json_response(ResponseHandler.update_failure("block"), 500)
#
#             return redirect('/add_block')
#
#         try:
#             # cursor.execute("SELECT District_id, District_Name FROM districts")
#             # districts = cursor.fetchall()
#
#             cursor.callproc("GetAllDistricts")
#             for dis in cursor.stored_results():
#                 districts = dis.fetchall()
#
#         except mysql.connector.Error as e:
#             print(f"Error fetching districts: {e}")
#             return json_response(ResponseHandler.fetch_failure("districts"), 500)
#
#         return render_template('edit_block.html', block_data=block_data, states=states, districts=districts)

@app.route('/edit_block/<int:block_id>', methods=['GET', 'POST'])
def edit_block(block_id):
    connection = config.get_db_connection()
    block_data = []
    states = []
    districts = []

    if connection:
        cursor = connection.cursor()
        # Retrieve all states
        try:
            cursor.execute("SELECT State_ID, State_Name FROM states")
            states = cursor.fetchall()
        except mysql.connector.Error as e:
            print(f"Error fetching states: {e}")
            return "Failed to fetch states", 500

        # Retrieve block data
        try:
            cursor.execute("SELECT Block_Name, District_id FROM blocks WHERE Block_Id = %s", (block_id,))
            block_data = cursor.fetchone()
        except mysql.connector.Error as e:
            print(f"Error fetching block data: {e}")
            return "Failed to fetch block data", 500

        # Handle POST request
        if request.method == 'POST':
            block_name = request.form['block_Name']
            district_id = request.form['district_Id']
            try:
                cursor.execute("UPDATE blocks SET Block_Name = %s, District_id = %s WHERE Block_Id = %s",
                               (block_name, district_id, block_id))
                connection.commit()
                flash("Block updated successfully!", "success")
                return redirect(url_for('edit_block', block_id=block_id))
            except mysql.connector.Error as e:
                print(f"Error updating blocks: {e}")
                return "Failed to update blocks", 500

        # Retrieve districts for the dropdown
        try:
            cursor.execute("SELECT District_id, District_Name FROM districts")
            districts = cursor.fetchall()
        except mysql.connector.Error as e:
            print(f"Error fetching districts: {e}")
            return "Failed to fetch districts", 500

        return render_template('edit_block.html', block_data=block_data, states=states, districts=districts)



# delete block by id
@app.route('/delete_block/<int:block_id>', methods=['GET', 'POST'])
def delete_block(block_id):
    connection = config.get_db_connection()

    if connection:
        cursor = connection.cursor()
        try:
            # cursor.execute("DELETE FROM blocks WHERE Block_Id = %s", (block_id,))
            cursor.callproc("DeleteBlock", (block_id,))
            connection.commit()
        except mysql.connector.Error as e:
            print(f"Error deleting block: {e}")
            return json_response(ResponseHandler.add_failure("block"), 500)
        finally:
            cursor.close()
            connection.close()

    return redirect('/add_block')


# this is get district all data by using state id ..
@app.route('/get_districts/<int:state_id>', methods=['GET'])
def get_districts(state_id):
    connection = config.get_db_connection()
    districts = []

    if connection:
        cursor = connection.cursor()
        try:
            # cursor.execute("SELECT District_id, District_Name FROM districts WHERE State_Id = %s", (state_id,))
            # districts = cursor.fetchall()

            cursor.callproc("GetDistrictsByStateId", (state_id,))
            for dis in cursor.stored_results():
                districts = dis.fetchall()
        except mysql.connector.Error as e:
            print(f"Error fetching districts: {e}")
            return json_response(ResponseHandler.fetch_failure("districts"), 500)
        finally:
            cursor.close()
            connection.close()

    return jsonify({
        "districts": [{"District_id": d[0], "District_Name": d[1]} for d in districts]
    })


# ----------- end Block controller -----------------

# ------------------------- Village controller ------------------------------------------
# Route to add a village
@app.route('/add_village', methods=['GET', 'POST'])
def add_village():
    connection = config.get_db_connection()
    cursor = connection.cursor()
    states = []
    villages = []

    try:
        # Fetch all states
        # cursor.execute("SELECT State_ID, State_Name FROM states")
        # states = cursor.fetchall()

        cursor.callproc("GetAllStates")
        for res in cursor.stored_results():
            states = res.fetchall()

        # Fetch all villages with their block names
        # cursor.execute("""
        #     SELECT v.Village_Id, v.Village_Name, b.Block_Name
        #     FROM villages v
        #     JOIN blocks b ON v.Block_Id = b.Block_Id
        # """)
        # villages = cursor.fetchall()
        cursor.callproc("GetAllVillages")
        for result in cursor.stored_results():
            villages = result.fetchall()

        if request.method == 'POST':
            block_id = request.form.get('block_Id')
            village_name = request.form.get('Village_Name', '').strip()

            if not block_id:
                return json_response(ResponseHandler.add_failure("block"), 400)

            if not re.match(str_pattern_reg, village_name):
                return json_response(ResponseHandler.invalid_name("village"), 400)

            # Check if the village already exists in the block
            cursor.execute("SELECT * FROM villages WHERE Village_Name = %s AND Block_Id = %s", (village_name, block_id))
            existing_village = cursor.fetchone()

            if existing_village:
                return json_response(ResponseHandler.already_exists("village"), 409)

            # Insert new village
            cursor.callproc('SaveVillage', (village_name, block_id))
            connection.commit()

            return json_response(ResponseHandler.add_success("village"), 200)

    except mysql.connector.Error as e:
        print(f"Database Error: {e}")
        return json_response(ResponseHandler.add_failure("village"), 500)
    finally:
        cursor.close()
        connection.close()

    return render_template('add_village.html', states=states, villages=villages)


# get block by district id
@app.route('/get_blocks/<int:district_id>', methods=['GET'])
def get_blocks(district_id):
    connection = config.get_db_connection()
    cursor = connection.cursor()
    blocks = []

    try:
        cursor.execute("SELECT Block_Id, Block_Name FROM blocks WHERE District_id = %s", (district_id,))
        blocks = cursor.fetchall()
    except mysql.connector.Error as e:
        print(f"Error fetching blocks: {e}")
        return json_response({"error": "Failed to fetch blocks"}, 500)
    finally:
        cursor.close()
        connection.close()

    return jsonify({"blocks": [{"Block_Id": block[0], "Block_Name": block[1]} for block in blocks]})


# check village
@app.route('/check_village', methods=['POST'])
def check_village():
    connection = config.get_db_connection()
    cursor = connection.cursor()

    block_id = request.form.get('block_Id')
    village_name = request.form.get('Village_Name', '').strip()

    # Validate village name
    if not re.match(str_pattern_reg, village_name):
        return json_response(ResponseHandler.invalid_name("village"), 400)

    if not block_id or not village_name:
        return json_response({'status': 'error', 'message': 'Block and Village Name are required!'}, 400)

    cursor.execute("SELECT * FROM villages WHERE Village_Name = %s AND Block_Id = %s", (village_name, block_id))
    existing_village = cursor.fetchone()

    cursor.close()
    connection.close()

    if existing_village:
        return json_response(ResponseHandler.already_exists("village"), 409)
    else:
        return json_response(ResponseHandler.is_available("village"), 200)


# update village
# @app.route('/edit_village/<int:village_id>', methods=['GET', 'POST'])
# def edit_village(village_id):
#     connection = config.get_db_connection()
#     village_data = None
#     blocks = []
#
#     try:
#         cursor = connection.cursor()
#         # Fetch village details
#         # cursor.execute("SELECT Village_Name, Block_Id FROM villages WHERE Village_Id = %s", (village_id,))
#         # village_data = cursor.fetchone()
#         cursor.callproc("GetVillageById", (village_id,))
#         for result in cursor.stored_results():
#             village_data = result.fetchone()
#
#         # Fetch all blocks for dropdown
#         # cursor.execute("SELECT Block_Id, Block_Name FROM blocks")
#         # blocks = cursor.fetchall()
#
#         cursor.callproc("GetAllBlocks")
#         for result in cursor.stored_results():
#             blocks = result.fetchall()
#
#         if request.method == 'POST':
#             village_name = request.form['Village_Name']
#             block_id = request.form['block_Id']
#
#             if not re.match(str_pattern_reg, village_name):
#                 return json_response(ResponseHandler.invalid_name("village"), 400)
#
#             # cursor.execute("UPDATE villages SET Village_Name = %s, Block_Id = %s WHERE Village_Id = %s",
#             #                (village_name, block_id, village_id))
#
#             cursor.callproc("UpdateVillage", (village_id, village_name, block_id,))
#
#             connection.commit()
#             return json_response(ResponseHandler.update_success("village"), 200)
#
#     except mysql.connector.Error as e:
#         print(f"Error: {e}")
#         return json_response(ResponseHandler.update_failure("village"), 500)
#     finally:
#         if cursor:
#             cursor.close()
#         if connection:
#             connection.close()
#
#     return render_template('edit_village.html', village_data=village_data, blocks=blocks)

@app.route('/edit_village/<int:village_id>', methods=['GET', 'POST'])
def edit_village(village_id):
    connection = config.get_db_connection()
    village_data = None
    blocks = []

    try:
        cursor = connection.cursor()
        # Fetch village details
        cursor.execute("SELECT Village_Name, Block_Id FROM villages WHERE Village_Id = %s", (village_id,))
        village_data = cursor.fetchone()

        # Fetch all blocks for dropdown
        cursor.execute("SELECT Block_Id, Block_Name FROM blocks")
        blocks = cursor.fetchall()

        if request.method == 'POST':
            village_name = request.form['Village_Name']
            block_id = request.form['block_Id']

            if not re.match("^[A-Za-z ]+$", village_name):
                flash("Invalid village name! Only letters and spaces allowed.", "error")
                return redirect(url_for('edit_village', village_id=village_id))

            cursor.execute("UPDATE villages SET Village_Name = %s, Block_Id = %s WHERE Village_Id = %s",
                           (village_name, block_id, village_id))
            connection.commit()
            flash("Village updated successfully!", "success")
            return redirect(url_for('edit_village', village_id=village_id))

    except mysql.connector.Error as e:
        print(f"Error: {e}")
        return "Failed to process request", 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

    return render_template('edit_village.html', village_data=village_data, blocks=blocks)


# delete village
@app.route('/delete_village/<int:village_id>', methods=['GET', 'POST'])
def delete_village(village_id):
    connection = config.get_db_connection()
    cursor = connection.cursor()

    try:
        # cursor.execute("DELETE FROM villages WHERE Village_Id = %s", (village_id,))
        cursor.callproc("DeleteVillage", (village_id,))
        connection.commit()
        # return json_response(ResponseHandler.delete_success("village"), 200)
    except mysql.connector.Error as e:
        print(f"Error: {e}")
        return json_response(ResponseHandler.add_failure("village"), 500)
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

    return redirect(url_for('add_village'))


# ---- end Village controller ---------------------


# -------------------------------- Invoice controller ------------------------------------------
@app.route('/add_invoice', methods=['GET', 'POST'])
def add_invoice():
    connection = config.get_db_connection()
    if not connection:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    if request.method == 'POST':
        try:
            cursor = connection.cursor(dictionary=True)

            # Get the village name from the form
            village_name = request.form.get('village')
            print("village name", village_name)

            # Query the database to get the corresponding Village_Id based on the village name
            cursor.execute("SELECT Village_Id FROM villages WHERE Village_Name = %s", (village_name,))
            village_result = cursor.fetchone()

            if not village_result:
                return jsonify({"status": "error", "message": f"Village '{village_name}' not found"}), 400

            village_id = village_result['Village_Id']

            # Fetch form data
            pmc_no = request.form.get('pmc_no')
            work_type = request.form.get('work_type')
            invoice_details = request.form.get('invoice_details')
            invoice_date = request.form.get('invoice_date')
            invoice_no = request.form.get('invoice_no')
            basic_amount = request.form.get('basic_amount')
            debit_amount = request.form.get('debit_amount')
            after_debit_amount = request.form.get('after_debit_amount')
            amount = request.form.get('amount')
            gst_amount = request.form.get('gst_amount')
            tds_amount = request.form.get('tds_amount')
            sd_amount = request.form.get('sd_amount')
            on_commission = request.form.get('on_commission')
            hydro_testing = request.form.get('hydro_testing')
            gst_sd_amount = request.form.get('gst_sd_amount')
            final_amount = request.form.get('final_amount')

            insert_invoice_query = '''
                INSERT INTO invoice (
                    PMC_No, Village_Id, Work_Type, Invoice_Details, Invoice_Date, Invoice_No, 
                    Basic_Amount, Debit_Amount, After_Debit_Amount, Amount, GST_Amount, TDS_Amount, 
                    SD_Amount, On_Commission, Hydro_Testing, GST_SD_Amount, Final_Amount
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            '''
            invoice_values = (
                pmc_no, village_id, work_type, invoice_details, invoice_date, invoice_no,
                basic_amount, debit_amount, after_debit_amount, amount, gst_amount, tds_amount,
                sd_amount, on_commission, hydro_testing, gst_sd_amount, final_amount
            )
            cursor.execute(insert_invoice_query, invoice_values)
            connection.commit()
            invoice_id = cursor.lastrowid

            # Insert into assign_subcontractors table
            subcontractor_id = request.form.get('subcontractor_id')
            insert_assign_query = '''
                INSERT INTO assign_subcontractors (PMC_no, Contractor_Id, Village_Id)
                VALUES (%s, %s, %s)
            '''
            cursor.execute(insert_assign_query, (pmc_no, subcontractor_id, village_id))
            connection.commit()

            # Insert Hold Amounts into invoice_subcontractor_hold_join table
            hold_types = request.form.getlist('hold_type[]')
            hold_amounts = request.form.getlist('hold_amount[]')
            hold_count = 0

            for hold_type, hold_amount in zip(hold_types, hold_amounts):
                cursor.execute("SELECT hold_type_id FROM hold_types WHERE hold_type = %s", (hold_type,))
                hold_type_result = cursor.fetchone()
                if not hold_type_result:
                    return jsonify({"status": "error", "message": f"Invalid Hold Type: {hold_type}"}), 400
                hold_type_id = hold_type_result['hold_type_id']
                insert_hold_query = '''
                    INSERT INTO invoice_subcontractor_hold_join (Contractor_Id, Invoice_Id, hold_type_id, hold_amount)
                    VALUES (%s, %s, %s, %s)
                '''
                cursor.execute(insert_hold_query, (subcontractor_id, invoice_id, hold_type_id, hold_amount))
                hold_count += 1

            connection.commit()

            return jsonify({"status": "success", "message": "Invoice added successfully"}), 201

        except mysql.connector.Error as e:
            connection.rollback()
            return jsonify({"status": "error", "message": f"Failed to add invoice: {str(e)}"}), 500
        finally:
            cursor.close()
            connection.close()

    # GET request: fetch and display all invoices (all fields) along with the form
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM view_invoice_details")
        invoices = cursor.fetchall()
        villages=[]
        cursor.callproc("GetAllVillages")
        for result in cursor.stored_results():
            villages = result.fetchall()

    except mysql.connector.Error as e:
        print(f"Error: {e}")
        invoices = []
    finally:
        cursor.close()
        connection.close()

    return render_template('add_invoice.html', invoices=invoices,villages=villages)

# search subcontraactor to assing invoice
@app.route('/search_subcontractor', methods=['POST'])
def search_subcontractor():
    connection = config.get_db_connection()
    if not connection:
        return json_response(ResponseHandler.fetch_failure("database connection"), 500)

    sub_query = request.form.get("query")
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute(
            "SELECT Contractor_Id, Contractor_Name FROM subcontractors WHERE Contractor_Name LIKE %s",
            (f"%{sub_query}%",)
        )
        results = cursor.fetchall()
        if not results:
            return "<li>No subcontractor found</li>"

        output = "".join(
            f"<li data-id='{row['Contractor_Id']}'>{row['Contractor_Name']}</li>"
            for row in results
        )
        return output

    except mysql.connector.Error as e:
        return json_response(ResponseHandler.fetch_failure(f"Search failed: {str(e)}"), 500)

    finally:
        cursor.close()
        connection.close()


# get hold types
@app.route('/get_hold_types', methods=['GET'])
def get_hold_types():
    connection = config.get_db_connection()
    try:
        cursor = connection.cursor(dictionary=True)
        # cursor.execute("SELECT hold_type_id, hold_type FROM hold_types")
        # hold_types = cursor.fetchall()

        cursor.callproc("GetAllHoldTypes")
        for hold in cursor.stored_results():
            hold_types = hold.fetchall()

        return jsonify(hold_types)
    except mysql.connector.Error as e:
        return ResponseHandler.fetch_failure({str(e)}), 500
        # return jsonify({"status": "error", "message": f"Failed to fetch hold types: {str(e)}"}), 500
    finally:
        cursor.close()
        connection.close()


# update invoice by id

@app.route('/edit_invoice/<int:invoice_id>', methods=['GET', 'POST'])
def edit_invoice(invoice_id):
    connection = config.get_db_connection()
    if not connection:
        return jsonify({"status": "error", "message": "Database connection failed"}), 500

    cursor = connection.cursor(dictionary=True)

    if request.method == 'POST':
        try:
            # Fetch updated form data
            subcontractor_id = request.form.get('subcontractor_id', '').strip()
            subcontractor_id = int(subcontractor_id) if subcontractor_id else None

            village_name = request.form.get('village')
            cursor.execute("SELECT Village_Id FROM villages WHERE Village_Name = %s", (village_name,))
            village_result = cursor.fetchone()
            if not village_result:
                return jsonify({"status": "error", "message": "Invalid Village Name"}), 400
            village_id = village_result['Village_Id']

            pmc_no = request.form.get('pmc_no')
            work_type = request.form.get('work_type')
            invoice_details = request.form.get('invoice_details')
            invoice_date = request.form.get('invoice_date')
            invoice_no = request.form.get('invoice_no')

            # Convert numeric fields properly
            numeric_fields = {
                "basic_amount": request.form.get('basic_amount'),
                "debit_amount": request.form.get('debit_amount'),
                "after_debit_amount": request.form.get('after_debit_amount'),
                "amount": request.form.get('amount'),
                "gst_amount": request.form.get('gst_amount'),
                "tds_amount": request.form.get('tds_amount'),
                "sd_amount": request.form.get('sd_amount'),
                "on_commission": request.form.get('on_commission'),
                "hydro_testing": request.form.get('hydro_testing'),
                "gst_sd_amount": request.form.get('gst_sd_amount'),
                "final_amount": request.form.get('final_amount'),
            }
            numeric_fields = {k: float(v) if v else 0 for k, v in numeric_fields.items()}

            # Update invoice
            update_invoice_query = '''
                UPDATE invoice
                SET PMC_No=%s, Village_Id=%s, Work_Type=%s, Invoice_Details=%s, Invoice_Date=%s,
                    Invoice_No=%s, Basic_Amount=%s, Debit_Amount=%s, After_Debit_Amount=%s, 
                    Amount=%s, GST_Amount=%s, TDS_Amount=%s, SD_Amount=%s, On_Commission=%s,
                    Hydro_Testing=%s, GST_SD_Amount=%s, Final_Amount=%s
                WHERE Invoice_Id=%s
            '''
            invoice_values = (
                pmc_no, village_id, work_type, invoice_details, invoice_date, invoice_no,
                *numeric_fields.values(), invoice_id
            )
            cursor.execute(update_invoice_query, invoice_values)
            connection.commit()

            # Handle holds
            hold_types = request.form.getlist('hold_type[]')
            hold_amounts = request.form.getlist('hold_amount[]')

            for hold_type, hold_amount in zip(hold_types, hold_amounts):
                if not hold_type:
                    continue  # skip empty hold types

                # Get or insert hold type
                cursor.execute("SELECT hold_type_id FROM hold_types WHERE hold_type = %s", (hold_type,))
                hold_type_result = cursor.fetchone()

                if not hold_type_result:
                    cursor.execute("INSERT INTO hold_types (hold_type) VALUES (%s)", (hold_type,))
                    connection.commit()
                    hold_type_id = cursor.lastrowid
                else:
                    hold_type_id = hold_type_result['hold_type_id']

                hold_amount = float(hold_amount) if hold_amount else 0

                # Check if join exists
                cursor.execute("""
                    SELECT join_id FROM invoice_subcontractor_hold_join
                    WHERE Invoice_Id = %s AND Contractor_Id = %s AND hold_type_id = %s
                """, (invoice_id, subcontractor_id, hold_type_id))
                join_result = cursor.fetchone()

                if join_result:
                    cursor.execute("""
                        UPDATE invoice_subcontractor_hold_join
                        SET hold_amount = %s
                        WHERE join_id = %s
                    """, (hold_amount, join_result['join_id']))
                else:
                    cursor.execute("""
                        INSERT INTO invoice_subcontractor_hold_join (Contractor_Id, Invoice_Id, hold_type_id, hold_amount)
                        VALUES (%s, %s, %s, %s)
                    """, (subcontractor_id, invoice_id, hold_type_id, hold_amount))

            connection.commit()
            return jsonify({"status": "success", "message": "Invoice updated successfully"}), 200

        except mysql.connector.Error as e:
            connection.rollback()
            return jsonify({"status": "error", "message": f"Failed to update invoice: {str(e)}"}), 500
        finally:
            cursor.close()
            connection.close()

    # ------------------ GET Request ------------------

    try:
        # Fetch invoice data
        cursor.execute(
            """SELECT i.*, s.Contractor_Name, v.Village_Name
               FROM invoice i
               LEFT JOIN assign_subcontractors a ON i.PMC_No = a.PMC_no AND i.Village_Id = a.Village_Id
               LEFT JOIN subcontractors s ON a.Contractor_Id = s.Contractor_Id
               LEFT JOIN villages v ON i.Village_Id = v.Village_Id
               WHERE i.Invoice_Id = %s""", (invoice_id,)
        )
        invoice = cursor.fetchone()
        if not invoice:
            return jsonify({"status": "error", "message": "Invoice not found"}), 404

        # Important! Clear unread result issue
        while cursor.nextset():
            pass

        # Fetch hold amounts
        cursor.execute(
            """SELECT h.hold_type, ihj.hold_amount 
               FROM invoice_subcontractor_hold_join ihj
               JOIN hold_types h ON ihj.hold_type_id = h.hold_type_id
               WHERE ihj.Invoice_Id = %s""", (invoice_id,)
        )
        hold_amounts = cursor.fetchall()
        invoice["hold_amounts"] = hold_amounts

    except mysql.connector.Error as e:
        return jsonify({"status": "error", "message": f"Database error: {str(e)}"}), 500
    finally:
        cursor.close()
        connection.close()

    return render_template('edit_invoice.html', invoice=invoice)



# delete invoice by id
@app.route('/delete_invoice/<int:invoice_id>', methods=['GET'])
def delete_invoice(invoice_id):
    connection = config.get_db_connection()
    if not connection:
        return json_response(ResponseHandler.fetch_failure("invoice"), 500)

    try:
        cursor = connection.cursor()
        # cursor.execute("DELETE FROM invoice WHERE Invoice_Id = %s", (invoice_id,))

        cursor.callproc("DeleteInvoice", (invoice_id,))
        connection.commit()

        # Check if the invoice was actually deleted
        if cursor.rowcount == 0:
            return json_response(ResponseHandler.fetch_failure("invoice"), 404)

        return redirect(url_for('add_invoice'))

    except mysql.connector.Error as e:
        print("Error deleting invoice:", e)
        return json_response(ResponseHandler.delete_failure("invoice"), 500)

    finally:
        cursor.close()
        connection.close()


# ---------- end Invoice controller ------------------


# ----------------------------- Payment controller ------------------------------------------
# this is Payment Page to add data
# @app.route('/add_payment', methods=['GET', 'POST'])
# def add_payment():
#     connection = config.get_db_connection()
#     payments = []  # List to hold payment history
#
#     if not connection:
#         return json_response(ResponseHandler.fetch_failure("payment"), 500)
#
#     try:
#         cursor = connection.cursor()
#
#         # Retrieve payment history
#         # cursor.execute(
#         #     "SELECT Payment_Id, PMC_No, Invoice_No, Payment_Amount, TDS_Payment_Amount, Total_Amount, UTR FROM payment"
#         # )
#         # payments = cursor.fetchall()
#         cursor.callproc("GetAllPayments")
#         for result in cursor.stored_results():
#             payments = result.fetchall()
#
#     except mysql.connector.Error as e:
#         print(f"Error fetching payment history: {e}")
#         return json_response(ResponseHandler.fetch_failure("payment"), 500)
#     finally:
#         cursor.close()
#
#     if request.method == 'POST':
#         pmc_no = request.form['PMC_No']
#         invoice_no = request.form['invoice_No']
#         amount = request.form['Payment_Amount']
#         tds_amount = request.form['TDS_Payment_Amount']
#         total_amount = request.form['total_amount']
#         utr = request.form['utr']
#
#         try:
#             cursor = connection.cursor()
#             cursor.callproc('SavePayment', (
#                 pmc_no, invoice_no, amount, tds_amount, total_amount, utr
#             ))
#             connection.commit()
#             return redirect(url_for('add_payment'))  # Redirect to add_payment page to reload the form
#         except mysql.connector.Error as e:
#             print(f"Error inserting payment: {e}")
#             return json_response(ResponseHandler.add_failure("payment"), 500)
#         finally:
#             cursor.close()
#             connection.close()
#
#     return render_template('add_payment.html', payments=payments)

@app.route('/add_payment', methods=['GET', 'POST'])
def add_payment():
    connection = config.get_db_connection()
    payments = []

    if connection:
        cursor = connection.cursor()

        try:
            cursor.execute(
                "SELECT Payment_Id, PMC_No, Invoice_No, Payment_Amount, TDS_Payment_Amount, Total_Amount, UTR FROM payment"
            )
            payments = cursor.fetchall()
        except mysql.connector.Error as e:
            print(f"Error fetching payment history: {e}")
            return "Failed to fetch payment history", 500
        finally:
            cursor.close()

    if request.method == 'POST':
        pmc_no = request.form['PMC_No']
        invoice_no = request.form['invoice_No']
        amount = request.form['Payment_Amount']
        tds_amount = request.form['TDS_Payment_Amount']
        total_amount = request.form['total_amount']
        utr = request.form['utr']

        try:
            cursor = connection.cursor()
            cursor.execute('''INSERT INTO payment (PMC_No, invoice_no, Payment_Amount, TDS_Payment_Amount, Total_Amount, UTR)
                              VALUES (%s, %s, %s, %s, %s, %s)''',
                           (pmc_no, invoice_no, amount, tds_amount, total_amount, utr))
            connection.commit()
            return redirect(url_for('add_payment'))

        except mysql.connector.Error as e:
            print(f"Error inserting payment: {e}")
            return "Failed to add payment", 500
        finally:
            cursor.close()
            connection.close()

    return render_template('add_payment.html', payments=payments)


@app.route('/get_pmc_nos_by_subcontractor/<subcontractorId>')
def get_pmc_nos_by_subcontractor(subcontractorId):
    connection = config.get_db_connection()
    cur = connection.cursor()
    print(subcontractorId)
    query = """
        SELECT DISTINCT i.PMC_No 
        FROM invoice i
        JOIN assign_subcontractors a ON i.PMC_No = a.PMC_no
        JOIN subcontractors s ON a.Contractor_Id = s.Contractor_Id
        WHERE s.Contractor_Id=%s;
    """
    cur.execute(query, (subcontractorId,))
    results = cur.fetchall()
    print(results)
    pmc_nos = [row[0] for row in results]
    cur.close()
    return jsonify({'pmc_nos': pmc_nos})



# Edit Payment Route
@app.route('/edit_payment/<int:payment_id>', methods=['GET', 'POST'])
def edit_payment(payment_id):
    connection = config.get_db_connection()
    payment_data = {}  # To hold the payment data for the given ID

    if not connection:
        return json_response(ResponseHandler.fetch_failure("payment"), 500)

    try:
        cursor = connection.cursor()

        # Fetch the existing payment data for the given payment_id
        # cursor.execute(
        #     "SELECT Payment_Id, PMC_No, Invoice_No, Payment_Amount, TDS_Payment_Amount, Total_Amount, UTR FROM payment WHERE Payment_Id = %s",
        #     (payment_id,)
        # )
        # payment_data = cursor.fetchone()

        cursor.callproc("GetPaymentById", (payment_id,))
        for result in cursor.stored_results():
            payment_data = result.fetchone()

        # Handle POST request to update the payment
        if request.method == 'POST':
            pmc_no = request.form['PMC_No']
            invoice_no = request.form['invoice_No']
            amount = request.form['Payment_Amount']
            tds_amount = request.form['TDS_Payment_Amount']
            total_amount = request.form['total_amount']
            utr = request.form['utr']

            try:
                # cursor.execute('''UPDATE payment SET PMC_No=%s, Invoice_No=%s, Payment_Amount=%s, TDS_Payment_Amount=%s,
                #                   Total_Amount=%s, UTR=%s WHERE Payment_Id=%s''',
                #                (pmc_no, invoice_no, amount, tds_amount, total_amount, utr, payment_id))

                cursor.callproc("UpdatePayment",(payment_id, pmc_no, invoice_no, amount, tds_amount, total_amount, utr,))
                connection.commit()

                return redirect(url_for('add_payment'))  # Redirect to add_payment page to view the updated list
            except mysql.connector.Error as e:
                print(f"Error updating payment: {e}")
                return json_response(ResponseHandler.update_failure("payment"), 500)

    except mysql.connector.Error as e:
        print(f"Error fetching payment data: {e}")
        return json_response(ResponseHandler.fetch_failure("payment"), 500)
    finally:
        cursor.close()
        connection.close()

    return render_template('edit_payment.html', payment_data=payment_data)


# Delete Payment Route
@app.route('/delete_payment/<int:payment_id>', methods=['GET', 'POST'])
def delete_payment(payment_id):
    connection = config.get_db_connection()
    if not connection:
        return json_response(ResponseHandler.fetch_failure("payment"), 500)
    try:
        cursor = connection.cursor()
        # cursor.execute("DELETE FROM payment WHERE Payment_Id = %s", (payment_id,))

        cursor.callproc("DeletePayment", (payment_id,))
        connection.commit()
        # Check if any rows were deleted
        if cursor.rowcount == 0:
            return json_response(ResponseHandler.fetch_failure("payment"), 404)
        return redirect(url_for('add_payment'))  # Redirect back to the add_payment page

    except mysql.connector.Error as e:
        print(f"Error deleting payment: {e}")
        return json_response(ResponseHandler.delete_failure("payment"), 500)
    finally:
        cursor.close()
        connection.close()


# --- end Payment controller -----------

# ------------------------- GST Release controller ------------------------------------------
@app.route('/add_gst_release', methods=['GET', 'POST'])
def add_gst_release():
    connection = config.get_db_connection()
    gst_releases = []  # List to hold GST Release history
    invoices = []  # List to hold invoices for the dropdown

    if not connection:
        return json_response(ResponseHandler.fetch_failure("GST Release"), 500)

    try:
        cursor = connection.cursor()

        # Retrieve GST Release history
        # cursor.execute("SELECT GST_Release_Id, PMC_No, Invoice_No, Basic_Amount, Final_Amount FROM gst_release")
        # gst_releases = cursor.fetchall()

        cursor.callproc("GetAllGSTReleases")
        for result in cursor.stored_results():
            gst_releases = result.fetchall()

        if request.method == 'POST':
            pmc_no = request.form['PMC_No']
            invoice_no = request.form['invoice_No']
            basic_amount = request.form['basic_amount']
            final_amount = request.form['final_amount']

            cursor.callproc('SaveGSTRelease', (
                pmc_no, invoice_no, basic_amount, final_amount
            ))
            connection.commit()

            return redirect(url_for('add_gst_release'))  # Redirect to add_gst_release page

    except mysql.connector.Error as e:
        print(f"Error: {e}")
        return json_response(ResponseHandler.add_failure("GST Release"), 500)

    finally:
        cursor.close()
        connection.close()

    return render_template('add_gst_release.html', invoices=invoices, gst_releases=gst_releases)


# update gst Release by id
@app.route('/edit_gst_release/<int:gst_release_id>', methods=['GET', 'POST'])
def edit_gst_release(gst_release_id):
    connection = config.get_db_connection()
    gst_release_data = {}  # To hold the GST release data for the given ID
    invoices = []  # List to hold invoices for the dropdown

    if not connection:
        return json_response(ResponseHandler.fetch_failure("GST Release"), 500)

    try:
        cursor = connection.cursor()

        # Fetch the existing GST release data for the given gst_release_id
        # cursor.execute(
        #     "SELECT GST_Release_Id, PMC_No, Invoice_No, Basic_Amount, Final_Amount FROM gst_release WHERE GST_Release_Id = %s",
        #     (gst_release_id,)
        # )
        # gst_release_data = cursor.fetchone()

        cursor.callproc("GetGSTReleaseById", (gst_release_id,))
        for result in cursor.stored_results():
            gst_release_data = result.fetchone()

        if request.method == 'POST':
            pmc_id = request.form['PMC_No']
            invoice_no = request.form['invoice_No']
            basic_amount = request.form['basic_amount']
            final_amount = request.form['final_amount']

            try:
                # cursor.execute('''UPDATE gst_release SET PMC_No=%s, Invoice_No=%s, Basic_Amount=%s, Final_Amount=%s
                #                   WHERE GST_Release_Id=%s''',
                #                (pmc_id, invoice_no, basic_amount, final_amount, gst_release_id))

                cursor.callproc("UpdateGSTRelease", (gst_release_id, pmc_id, invoice_no, basic_amount, final_amount ))

                connection.commit()

                return redirect(url_for('add_gst_release'))  # Redirect to the page to view the updated list

            except mysql.connector.Error as e:
                print(f"Error updating GST Release: {e}")
                return json_response(ResponseHandler.update_failure("GST Release"), 500)

    except mysql.connector.Error as e:
        print(f"Error fetching GST Release data: {e}")
        return json_response(ResponseHandler.fetch_failure("GST Release"), 500)

    finally:
        cursor.close()
        connection.close()

    return render_template('edit_gst_release.html', gst_release_data=gst_release_data, invoices=invoices)


# delete gst release by id
@app.route('/delete_gst_release/<int:gst_release_id>', methods=['GET', 'POST'])
def delete_gst_release(gst_release_id):
    connection = config.get_db_connection()
    if not connection:
        return json_response(ResponseHandler.fetch_failure("GST Release"), 500)
    try:
        cursor = connection.cursor()
        # cursor.execute("DELETE FROM gst_release WHERE GST_Release_Id = %s", (gst_release_id,))
        cursor.callproc("DeleteGSTRelease", (gst_release_id,))
        connection.commit()
        # Check if any rows were deleted
        if cursor.rowcount == 0:
            return json_response(ResponseHandler.fetch_failure("GST Release"), 404)
        return redirect(url_for('add_gst_release'))  # Redirect to the add_gst_release page
    except mysql.connector.Error as e:
        print(f"Error deleting GST Release: {e}")
        return json_response(ResponseHandler.delete_failure("GST Release"), 500)
    finally:
        cursor.close()
        connection.close()


# --- end GST Release controller  -----

# ------------------------- Subcontractor controller ------------------------------------------
@app.route('/subcontractor', methods=['GET', 'POST'])
def subcontract():
    connection = config.get_db_connection()
    subcontractor = []

    if not connection:
        return json_response(ResponseHandler.fetch_failure("Subcontractor"), 500)

    try:
        cursor = connection.cursor()

        if request.method == 'GET':
            try:
                cursor.execute('SELECT * FROM subcontractors;')
                subcontractor = cursor.fetchall()  # Fetch the current subcontractor list
                connection.commit()
            except Error as e:
                print(f"Error fetching data: {e}")
                return json_response(ResponseHandler.fetch_failure("Subcontractor"), 500)

        if request.method == 'POST':
            contractor_data = {
                'Contractor_Name': request.form['Contractor_Name'],
                'Address': request.form['Address'],
                'Mobile_No': request.form['Mobile_No'],
                'PAN_No': request.form['PAN_No'],
                'Email': request.form['Email'],
                'Gender': request.form['Gender'],
                'GST_Registration_Type': request.form['GST_Registration_Type'],
                'GST_No': request.form['GST_No'],
                'Contractor_password': request.form['Contractor_password'],
            }

            try:
                cursor.callproc('SaveContractor', (
                    contractor_data['Contractor_Name'],
                    contractor_data['Address'],
                    contractor_data['Mobile_No'],
                    contractor_data['PAN_No'],
                    contractor_data['Email'],
                    contractor_data['Gender'],
                    contractor_data['GST_Registration_Type'],
                    contractor_data['GST_No'],
                    contractor_data['Contractor_password']
                ))
                connection.commit()

                # Re-fetch subcontractors after inserting the new one
                cursor.execute('SELECT * FROM subcontractors')
                subcontractor = cursor.fetchall()

            except Error as e:
                print(f"Error inserting data: {e}")
                return json_response(ResponseHandler.add_failure("Subcontractor"), 500)

    except Error as e:
        print(f"Error handling subcontractor data: {e}")
        return json_response(ResponseHandler.fetch_failure("Subcontractor"), 500)

    finally:
        cursor.close()
        connection.close()

    return render_template('add_subcontractor.html', subcontractor=subcontractor)


# update subcontractor by id
@app.route('/edit_subcontractor/<int:id>', methods=['GET', 'POST'])
def edit_subcontractor(id):
    connection = config.get_db_connection()

    if not connection:
        return json_response(ResponseHandler.fetch_failure("Subcontractor"), 500)

    try:
        cursor = connection.cursor()
        subcontractor = None

        # Fetch existing subcontractor data by ID
        # cursor.execute('SELECT * FROM subcontractors WHERE Contractor_Id = %s', (id,))
        # subcontractor = cursor.fetchone()

        cursor.callproc("GetSubcontractorById", (id,))
        for contractors in cursor.stored_results():
            subcontractor = contractors.fetchone()

        if not subcontractor:
            return json_response(ResponseHandler.fetch_failure("Subcontractor"), 404)

        if request.method == 'POST':
            updated_data = {
                'Contractor_Name': request.form['Contractor_Name'],
                'Address': request.form['Address'],
                'Mobile_No': request.form['Mobile_No'],
                'PAN_No': request.form['PAN_No'],
                'Email': request.form['Email'],
                'Gender': request.form['Gender'],
                'GST_Registration_Type': request.form['GST_Registration_Type'],
                'GST_No': request.form['GST_No'],
                'Contractor_password': request.form['Contractor_password'],
                'id': id
            }

            try:
                # cursor.execute("""UPDATE subcontractors SET
                #     Contractor_Name=%(Contractor_Name)s,
                #     Address=%(Address)s,
                #     Mobile_No=%(Mobile_No)s,
                #     PAN_No=%(PAN_No)s,
                #     Email=%(Email)s,
                #     Gender=%(Gender)s,
                #     GST_Registration_Type=%(GST_Registration_Type)s,
                #     GST_No=%(GST_No)s,
                #     Contractor_password=%(Contractor_password)s
                #     WHERE Contractor_Id=%(id)s""", updated_data)

                cursor.callproc("UpdateSubcontractor", (
                    id,
                    updated_data['Contractor_Name'],
                    updated_data['Address'],
                    updated_data['Mobile_No'],
                    updated_data['PAN_No'],
                    updated_data['Email'],
                    updated_data['Gender'],
                    updated_data['GST_Registration_Type'],
                    updated_data['GST_No'],
                    updated_data['Contractor_password']
                ))
                connection.commit()
                return redirect(url_for('subcontract'))

            except Error as e:
                print(f"Error updating subcontractor: {e}")
                return json_response(ResponseHandler.update_failure("Subcontractor"), 500)

    except Error as e:
        print(f"Error fetching subcontractor data: {e}")
        return json_response(ResponseHandler.fetch_failure("Subcontractor"), 500)

    finally:
        cursor.close()
        connection.close()

    return render_template('edit_subcontractor.html', subcontractor=subcontractor)


# delete Sub-Contractor methods by id ..
@app.route('/deleteSubContractor/<int:id>', methods=['GET', 'POST'])
def deleteSubContractor(id):
    connection = config.get_db_connection()

    if not connection:
        return json_response(ResponseHandler.fetch_failure("Subcontractor"), 500)

    try:
        cursor = connection.cursor()

        # cursor.execute("DELETE FROM subcontractors WHERE Contractor_Id = %s", (id,))
        cursor.callproc("DeleteSubcontractor", (id,))
        connection.commit()

        # Check if any row was deleted (subcontractor found)
        if cursor.rowcount == 0:
            return json_response(ResponseHandler.fetch_failure("Subcontractor"), 404)

    except Error as e:
        print(f"Error deleting subcontractor: {e}")
        return json_response(ResponseHandler.delete_failure("Subcontractor"), 500)

    finally:
        cursor.close()
        connection.close()

    return redirect(url_for('subcontract'))


# ------------------------------- Show Report Subcontractor  ---------------------

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)


# Upload Excel file html page
@app.route('/upload_excel_file', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        file = request.files['file']
        if file and file.filename.endswith('.xlsx'):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(filepath)
            return redirect(url_for('show_table', filename=file.filename))
    return render_template('uploadExcelFile.html')


# Show excel data in tables6
@app.route('/show_table/<filename>')
def show_table(filename):
    global data
    data = []

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active

    # Extract key file information from the first 4 rows
    file_info = {
        "Subcontractor": sheet.cell(row=1, column=2).value,
        "State": sheet.cell(row=2, column=2).value,
        "District": sheet.cell(row=3, column=2).value,
        "Block": sheet.cell(row=4, column=2).value,
    }

    errors = []
    subcontractor_data = None
    state_data = None
    district_data = None
    block_data = None

    # Database connection
    connection = config.get_db_connection()
    if connection:
        try:
            cursor = connection.cursor(dictionary=True)

            # Validate State
            cursor.execute("SELECT State_ID, State_Name FROM states WHERE State_Name = %s", (file_info['State'],))
            state_data = cursor.fetchone()
            if not state_data:
                errors.append(f"State '{file_info['State']}' is not valid. Please add it.")

            # Validate District
            if state_data:
                cursor.execute(
                    "SELECT District_ID, District_Name FROM districts WHERE District_Name = %s AND State_ID = %s",
                    (file_info['District'], state_data['State_ID'])
                )
                district_data = cursor.fetchone()
                if not district_data:
                    errors.append(
                        f"District '{file_info['District']}' is not valid under state '{file_info['State']}'.")

            # Validate Block
            if district_data:
                cursor.execute(
                    "SELECT Block_Id, Block_Name FROM blocks WHERE Block_Name = %s AND District_ID = %s",
                    (file_info['Block'], district_data['District_ID'])
                )
                block_data = cursor.fetchone()
                if not block_data:
                    errors.append(
                        f"Block '{file_info['Block']}' is not valid under district '{file_info['District']}'.")


            # old code
            # # Validate Subcontractor
            cursor.execute("SELECT Contractor_Id, Contractor_Name FROM subcontractors WHERE Contractor_Name = %s",
                           (file_info['Subcontractor'],))
            subcontractor_data = cursor.fetchone()

            if not subcontractor_data:
                cursor.execute("INSERT INTO subcontractors (Contractor_Name) VALUES (%s)",
                               (file_info['Subcontractor'],))
                connection.commit()
                cursor.execute("SELECT Contractor_Id, Contractor_Name FROM subcontractors WHERE Contractor_Name = %s",
                               (file_info['Subcontractor'],))
                subcontractor_data = cursor.fetchone()

            # new code
            # cursor.callproc('ValidateAndInsertSubcontractor', (file_info['Subcontractor'], 0, ''))
            #
            # for con in cursor.stored_results():
            #     subcontractor_data = con.fetchone()
            #     print("subcon:",subcontractor_data)
            #
            # print("subcontractor_data",subcontractor_data)

            # Get hold types data from database (for faster lookup)
            # cursor.execute("SELECT hold_type_id, hold_type FROM hold_types")
            # hold_types_data = cursor.fetchall()

            cursor.callproc("GetAllHoldTypes")
            for ht in cursor.stored_results():
                hold_types_data = ht.fetchall()


            hold_types_lookup = {row['hold_type'].lower(): row['hold_type_id'] for row in hold_types_data if row['hold_type']}


            cursor.close()
        except mysql.connector.Error as e:
            print(f"Database error: {e}")

           # return "Database operation failed", 500
            return f"{e}",500
        finally:
            connection.close()

    # Extract dynamic variable names from row 5 and detect "hold" columns
    variables = {}
    hold_columns = []
    hold_counter = 0

    for j in range(1, sheet.max_column + 1):
        col_value = sheet.cell(row=5, column=j).value
        if col_value:
            variables[col_value] = j  # Store column name with its position

            # Check if the column header contains the word 'hold'
            if 'hold' in str(col_value).lower():
                hold_counter += 1
                # Lookup hold type id from database
                hold_type_key = str(col_value).lower().strip()
                hold_type_id = hold_types_lookup.get(hold_type_key, None)
                hold_columns.append({
                    'column_name': col_value,
                    'column_number': j,
                    'hold_type_id': hold_type_id
                })

    # Extract data dynamically based on row numbers
    for i in range(6, sheet.max_row + 1):
        row_data = {}
        if sheet.cell(row=i, column=1).value:
            row_data["Row Number"] = i  # Store row number
            for var_name, col_num in variables.items():
                row_data[var_name] = sheet.cell(row=i, column=col_num).value
            # Check if at least 4 non-empty cells exist in the row
            if sum(1 for value in row_data.values() if value) >= 4:
                data.append(row_data)

    # For debugging or console output, you can print the hold columns info
    for hold in hold_columns:
        if hold['hold_type_id']:
            print(
                f" if Column: {hold['column_name']}, Column Number: {hold['column_number']}, Hold Type ID: {hold['hold_type_id']}")
        else:
            errors.append(
                f"Hold Type not added ! Column name '{hold['column_name']}'.")
            print(
                f" else Column: {hold['column_name']}, Column Number: {hold['column_number']}, Hold Type ID: {hold['hold_type_id']}")

    return render_template(
        'show_excel_file.html',
        file_info=file_info,
        variables=variables,
        data=data,
        subcontractor_data=subcontractor_data,
        state_data=state_data,
        district_data=district_data,
        block_data=block_data,
        errors=errors,
        hold_columns=hold_columns,
        hold_counter=hold_counter
    )


# save Excel data
@app.route('/save_data', methods=['POST'])
def save_data():
    # Extract form data
    subcontractor_id = request.form.get("subcontractor_data")
    state_id = request.form.get("state_data")
    district_id = request.form.get("district_data")
    block_id = request.form.get("block_data")

    variables = request.form.getlist('variables[]')
    hold_columns = request.form.get("hold_columns")
    hold_counter = request.form.get("hold_counter")

    # print("Info: ", subcontractor_id, state_id, district_id, block_id)

    if not data:
        return jsonify({"error": "No data provided to save"}), 400

    if data:
        # print("Total number of entries in data:", len(data))

        connection = config.get_db_connection()
        cursor = connection.cursor()

        try:
            for entry in data:
                save_data = {
                    "PMC_No": entry.get("PMC_No"),
                    "Invoice_Details": entry.get("Invoice_Details", ''),
                    "Work_Type": 'none',
                    "Invoice_Date": entry.get("Invoice_Date").strftime('%Y-%m-%d') if entry.get(
                        "Invoice_Date") else None,
                    "Invoice_No": entry.get("Invoice_No", ''),
                    "Basic_Amount": entry.get("Basic_Amount", 0.00),
                    "Debit_Amount": entry.get("Debit_Amount", 0.00),
                    "After_Debit_Amount": entry.get("After_Debit_Amount", 0.00),
                    "Amount": entry.get("Amount", 0.00),
                    "GST_Amount": entry.get("GST_Amount", 0.00),
                    "TDS_Amount": entry.get("TDS_Amount", 0.00),
                    "SD_Amount": entry.get("SD_Amount", 0.00),
                    "On_Commission": entry.get("On_Commission", 0.00),
                    "Hydro_Testing": entry.get("Hydro_Testing", 0.00),
                    "Hold_Amount": 0,
                    "GST_SD_Amount": entry.get("GST_SD_Amount", 0.00),
                    "Final_Amount": entry.get("Final_Amount", 0.00),
                    "Payment_Amount": entry.get("Payment_Amount", 0.00),
                    "Total_Amount": entry.get("Total_Amount", 0.00),
                    "TDS_Payment_Amount": entry.get("TDS_Payment_Amount", 0.00),
                    "UTR": entry.get("UTR", ''),
                }

                village_name, work_type = None, None
                village_id = 0

                PMC_No = save_data.get('PMC_No')
                Invoice_Details = save_data.get('Invoice_Details')
                Invoice_Date = save_data.get('Invoice_Date')
                Invoice_No = save_data.get('Invoice_No')
                Basic_Amount = save_data.get('Basic_Amount')
                Debit_Amount = save_data.get('Debit_Amount')
                After_Debit_Amount = save_data.get('After_Debit_Amount')
                Amount = save_data.get('Amount')
                GST_Amount = save_data.get('GST_Amount')
                TDS_Amount = save_data.get('TDS_Amount')
                SD_Amount = save_data.get('SD_Amount')
                On_Commission = save_data.get('On_Commission')
                Hydro_Testing = save_data.get('Hydro_Testing')
                GST_SD_Amount = save_data.get('GST_SD_Amount')
                Final_Amount = save_data.get('Final_Amount')

                Payment_Amount = save_data.get('Payment_Amount')
                Total_Amount = save_data.get('Total_Amount')
                TDS_Payment_Amount = save_data.get('TDS_Payment_Amount')
                UTR = save_data.get('UTR')


                if Invoice_Details:
                    words = Invoice_Details.lower().split()
                    if 'village' in words:
                        village_pos = words.index('village')
                        village_name = " ".join(words[:village_pos])
                    if 'work' in words:
                        work_pos = words.index('work')
                        if village_name:
                            work_type = " ".join(words[village_pos + 1:work_pos + 1])
                        else:
                            work_type = " ".join(words[:work_pos + 1])

                    if Invoice_Details and 'village' in Invoice_Details.lower() and 'work' in Invoice_Details.lower():
                        print("village_name ::", village_name, "|| work_type ::", work_type)
                        if block_id and village_name:
                            village_id = None
                            # cursor.execute("SELECT Village_Id FROM villages WHERE Block_Id = %s AND Village_Name = %s",(block_id, village_name))
                            # result = cursor.fetchone()
                            cursor.callproc("GetVillageId",(block_id, village_name))
                            for result in cursor.stored_results():
                                result = result.fetchone()
                            village_id = result[0] if result else None

                            if not village_id:
                                # cursor.execute("INSERT INTO villages (Village_Name, Block_Id) VALUES (%s, %s)", (village_name, block_id))

                                cursor.callproc("SaveVillage", (village_name, block_id))
                                # cursor.execute("SELECT Village_Id FROM villages WHERE Block_Id = %s AND Village_Name = %s",(block_id, village_name))
                                # result = cursor.fetchone()
                                cursor.callproc("GetVillageId", (block_id, village_name))
                                for result in cursor.stored_results():
                                    result = result.fetchone()
                                village_id = result[0] if result else None

                        print("village_id :", village_id)
                        print("block_id :", block_id)
                        print("invoice :", PMC_No, village_id, work_type, Invoice_Details, Invoice_Date, Invoice_No,
                              Basic_Amount, Debit_Amount, After_Debit_Amount, Amount, GST_Amount, TDS_Amount,
                              SD_Amount, On_Commission, Hydro_Testing, GST_SD_Amount, Final_Amount)
                        #
                        # cursor.execute("SET @p_invoice_id = 0")
                        # cursor.callproc("SaveInvoice", (
                        #     PMC_No, village_id, work_type, Invoice_Details, Invoice_Date, Invoice_No,
                        #     Basic_Amount, Debit_Amount, After_Debit_Amount, Amount, GST_Amount, TDS_Amount,
                        #     SD_Amount, On_Commission, Hydro_Testing, GST_SD_Amount, Final_Amount,
                        #     subcontractor_id, "@p_invoice_id"
                        # ))
                        # cursor.execute("SELECT @p_invoice_id")
                        # invoice_id = cursor.fetchone()[0]

                        args = (
                            PMC_No, village_id, work_type, Invoice_Details, Invoice_Date, Invoice_No,
                            Basic_Amount, Debit_Amount, After_Debit_Amount, Amount, GST_Amount, TDS_Amount,
                            SD_Amount, On_Commission, Hydro_Testing, GST_SD_Amount, Final_Amount,
                            subcontractor_id, 0
                        )

                        results = cursor.callproc('SaveInvoice', args)
                        invoice_id = results[-1]

                        print("invoice id ",invoice_id)


                        if isinstance(hold_columns, str):
                            hold_columns = ast.literal_eval(hold_columns)

                            # Check if hold_columns is actually a list of dictionaries
                        if isinstance(hold_columns, list) and all(isinstance(hold, dict) for hold in hold_columns):
                            for hold in hold_columns:
                                print(f"Processing hold: {hold}")
                                hold_column_name = hold.get('column_name')  # Get column name
                                hold_type_id = hold.get('hold_type_id')  # Get hold_type_id

                                if hold_column_name:
                                    hold_amount = entry.get(
                                        hold_column_name)  # Get the value for that specific hold column
                                    if hold_amount is not None:
                                        print(f"Processing hold type: {hold_column_name}, Hold Amount: {hold_amount}")

                                        # Insert into the invoice_subcontractor_hold_join table
                                        hold_join_data = {
                                            "Contractor_Id": subcontractor_id,
                                            "Invoice_Id": invoice_id,
                                            "hold_type_id": hold_type_id,
                                            "hold_amount": hold_amount
                                        }

                                        insert_hold_query = """INSERT INTO invoice_subcontractor_hold_join (Contractor_Id, Invoice_Id, hold_type_id, hold_amount)
                                                            VALUES (%(Contractor_Id)s, %(Invoice_Id)s, %(hold_type_id)s, %(hold_amount)s);
                                                        """
                                        cursor.execute(insert_hold_query, hold_join_data)
                                        print(f"Inserted hold join data: {hold_join_data}")

                                else:
                                    print(f"Invalid hold entry: {hold}")
                        else:
                            print("Hold columns data is not a valid list of dictionaries.")


                    elif Invoice_Details and any(
                            keyword in Invoice_Details.lower() for keyword in ['gst', 'release', 'note']):
                        print("Gst rels :", PMC_No, Invoice_No, Basic_Amount, Final_Amount)
                        cursor.callproc("SaveGSTRelease",(PMC_No, Invoice_No, Basic_Amount, Final_Amount))
                        # cursor.execute(
                        #     """INSERT INTO gst_release (PMC_No, Invoice_No, Basic_Amount, Final_Amount) VALUES (%s,%s, %s, %s)""",
                        #     (PMC_No, Invoice_No, Basic_Amount, Final_Amount))

                        # insert_payment = """INSERT INTO payment (PMC_No, Invoice_No, Payment_Amount, TDS_Payment_Amount, Total_Amount, UTR) VALUES (%s, %s, %s, %s, %s, %s)"""
                        # cursor.execute(insert_payment,
                        #                (PMC_No, Invoice_No, Payment_Amount, TDS_Payment_Amount, Total_Amount, UTR))

                if PMC_No and Total_Amount and UTR:
                    print("Payment :", PMC_No, Invoice_No, Payment_Amount, TDS_Payment_Amount, Total_Amount, UTR)
                    # insert_payment = """INSERT INTO payment (PMC_No, Invoice_No, Payment_Amount, TDS_Payment_Amount, Total_Amount, UTR) VALUES (%s, %s, %s, %s, %s, %s)"""
                    # cursor.execute(insert_payment,
                    #                (PMC_No, Invoice_No, Payment_Amount, TDS_Payment_Amount, Total_Amount, UTR))

                    cursor.callproc("SavePayment",(PMC_No, Invoice_No, Payment_Amount, TDS_Payment_Amount, Total_Amount, UTR))

            connection.commit()
            return jsonify({"success": "Data saved successfully!"}), 200
            # return render_template('uploadExcelFile.html')
        except Exception as e:
            connection.rollback()
            return jsonify({"error": f"An unexpected error occurred: {e}"}), 500
        finally:
            cursor.close()
            connection.close()

    return render_template('index.html')


# ---------------------- Report --------------------------------
# call repor page
@app.route('/report')
def report_page():
    return render_template('report.html')


# Search list multiples input and search reports
@app.route('/search_contractor', methods=['POST'])
def search_contractor():
    connection = config.get_db_connection()
    cursor = connection.cursor(dictionary=True)

    subcontractor_name = request.form.get('subcontractor_name')
    pmc_no = request.form.get('pmc_no')
    state = request.form.get('state')
    district = request.form.get('district')
    block = request.form.get('block')
    village = request.form.get('village')
    year_from = request.form.get('year_from')
    year_to = request.form.get('year_to')

    conditions = []
    params = []

    if subcontractor_name:
        conditions.append("LOWER(s.Contractor_Name) LIKE LOWER(%s)")
        params.append(f"%{subcontractor_name}%")
    if pmc_no:
        conditions.append("i.PMC_No = %s")
        params.append(pmc_no)
    if state:
        conditions.append("LOWER(st.State_Name) LIKE LOWER(%s)")
        params.append(f"%{state}%")
    if district:
        conditions.append("LOWER(d.District_Name) LIKE LOWER(%s)")
        params.append(f"%{district}%")
    if block:
        conditions.append("LOWER(b.Block_Name) LIKE LOWER(%s)")
        params.append(f"%{block}%")
    if village:
        conditions.append("LOWER(v.Village_Name) LIKE LOWER(%s)")
        params.append(f"%{village}%")
    if year_from and year_to:
        conditions.append("i.Invoice_Date BETWEEN %s AND %s")
        params.append(year_from)
        params.append(year_to)

    if not conditions:
        return jsonify({"error": "At least one field is required for search."}), 400

    query = f""" SELECT DISTINCT s.Contractor_Id, s.Contractor_Name, i.PMC_No, st.State_Name, 
        d.District_Name, b.Block_Name, v.Village_Name 
        FROM subcontractors s
        LEFT JOIN assign_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
        LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
        LEFT JOIN blocks b ON v.Block_Id = b.Block_Id
        LEFT JOIN districts d ON b.District_id = d.District_id
        LEFT JOIN states st ON d.State_Id = st.State_Id
        LEFT JOIN invoice i ON v.Village_Id = i.Village_Id
        WHERE {' AND '.join(conditions)}
    """

    cursor.execute(query, tuple(params))
    data = cursor.fetchall()
    return jsonify(data)


# @app.route('/contractor_report/<int:contractor_id>')
# def contractor_report(contractor_id):
#     connection = config.get_db_connection()
#     cursor = connection.cursor(dictionary=True, buffered=True)
#
#     try:
#         # Fetch contractor details
#         cursor.execute("""
#             SELECT DISTINCT s.Contractor_Name, st.State_Name, d.District_Name, b.Block_Name,
#                    s.Mobile_No, s.GST_Registration_Type, s.GST_No,s.PAN_No,s.Email,s.Address
#             FROM subcontractors s
#             LEFT JOIN assign_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
#             LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
#             LEFT JOIN blocks b ON v.Block_Id = b.Block_Id
#             LEFT JOIN districts d ON b.District_id = d.District_id
#             LEFT JOIN states st ON d.State_Id = st.State_Id
#             WHERE s.Contractor_Id = %s
#         """, (contractor_id,))
#         contInfo = cursor.fetchone()
#
#         # cursor.callproc('GetContractorInfoById', [contractor_id])
#         # contInfo = next(cursor.stored_results()).fetchone()
#
#         # Fetch distinct hold types present in invoices for the contractor
#         cursor.execute("""
#             SELECT DISTINCT ht.hold_type_id, ht.hold_type
#             FROM invoice_subcontractor_hold_join h
#             JOIN hold_types ht ON h.hold_type_id = ht.hold_type_id
#             JOIN invoice i ON h.Invoice_Id = i.Invoice_Id
#             WHERE h.Contractor_Id = %s
#         """, (contractor_id,))
#         hold_types = cursor.fetchall()
#         # cursor.callproc('GetHoldTypesByContId', [contractor_id])
#         # hold_types = next(cursor.stored_results()).fetchall()
#
#         # Fetch all invoices for the contractor, with optional hold information
#         query = """
#             SELECT DISTINCT i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details,
#                    i.Invoice_Date, i.Invoice_No, i.Basic_Amount, i.Debit_Amount,
#                    i.After_Debit_Amount, i.Amount, i.GST_Amount, i.TDS_Amount, i.SD_Amount,
#                    i.On_Commission, i.Hydro_Testing, i.GST_SD_Amount,
#                    i.Final_Amount, h.hold_amount, ht.hold_type
#             FROM invoice i
#             LEFT JOIN villages v ON i.Village_Id = v.Village_Id
#             LEFT JOIN assign_subcontractors asg ON v.Village_Id = asg.Village_Id
#             LEFT JOIN subcontractors s ON asg.Contractor_Id = s.Contractor_Id
#             LEFT JOIN invoice_subcontractor_hold_join h ON i.Invoice_Id = h.Invoice_Id AND h.Contractor_Id = s.Contractor_Id
#             LEFT JOIN hold_types ht ON h.hold_type_id = ht.hold_type_id
#             WHERE s.Contractor_Id = %s
#             ORDER BY i.PMC_No ASC;
#         """
#         cursor.execute(query, (contractor_id,))
#         invoices = cursor.fetchall()
#
#         # cursor.callproc('GetInvoicesWithHoldByContId', [contractor_id])
#         # invoices = next(cursor.stored_results()).fetchall()
#
#         gst_query = """select pmc_no,invoice_no,basic_amount,final_amount from gst_release where pmc_no in
#                         (select distinct pmc_no from assign_subcontractors where Contractor_Id= %s ) ORDER BY pmc_no ASC """
#         cursor.execute(gst_query, (contractor_id,))
#         gst_rel = cursor.fetchall()
#
#         # cursor.callproc('GetGstReleaseBycontId', [contractor_id])
#         # gst_rel = next(cursor.stored_results()).fetchall()
#
#         pay_query = """select  pmc_no,invoice_no,Payment_Amount,TDS_Payment_Amount,Total_amount,utr from payment where pmc_no in
#                         (select distinct pmc_no from assign_subcontractors where Contractor_Id=%s ) ORDER BY pmc_no ASC """
#         cursor.execute(pay_query, (contractor_id,))
#         payments = cursor.fetchall()
#         # cursor.callproc('GetPaymentsByContId', [contractor_id])
#         # payments = next(cursor.stored_results()).fetchall()
#
#         total = {
#             "sum_invo_basic_amt": float(sum(row['Basic_Amount'] or 0 for row in invoices)),
#             "sum_invo_debit_amt": float(sum(row['Debit_Amount'] or 0 for row in invoices)),
#             "sum_invo_after_debit_amt": float(sum(row['After_Debit_Amount'] or 0 for row in invoices)),
#             "sum_invo_amt": float(sum(row['Amount'] or 0 for row in invoices)),
#             "sum_invo_gst_amt": float(sum(row['GST_Amount'] or 0 for row in invoices)),
#             "sum_invo_tds_amt": float(sum(row['TDS_Amount'] or 0 for row in invoices)),
#             "sum_invo_ds_amt": float(sum(row['SD_Amount'] or 0 for row in invoices)),
#             "sum_invo_on_commission": float(sum(row['On_Commission'] or 0 for row in invoices)),
#             "sum_invo_hydro_test": float(sum(row['Hydro_Testing'] or 0 for row in invoices)),
#             "sum_invo_gst_sd_amt": float(sum(row['GST_SD_Amount'] or 0 for row in invoices)),
#             "sum_invo_final_amt": float(sum(row['Final_Amount'] or 0 for row in invoices)),
#             "sum_invo_hold_amt": float(sum(row['hold_amount'] or 0 for row in invoices)),
#
#             "sum_gst_basic_amt": float(sum(row['basic_amount'] or 0 for row in gst_rel)),
#             "sum_gst_final_amt": float(sum(row['final_amount'] or 0 for row in gst_rel)),
#
#             "sum_pay_payment_amt": float(sum(row['Payment_Amount'] or 0 for row in payments)),
#             "sum_pay_tds_payment_amt": float(sum(row['TDS_Payment_Amount'] or 0 for row in payments)),
#             "sum_pay_total_amt": float(sum(row['Total_amount'] or 0 for row in payments))
#         }
#         current_date = datetime.now().strftime('%Y-%m-%d')
#
#     except Exception as e:
#         print(f"Error fetching contractor report: {e}")
#         return "An error occurred while fetching contractor report", 500
#
#     finally:
#         cursor.close()
#         connection.close()
#
#     return render_template('subcontractor_report.html', contInfo=contInfo, contractor_id=contractor_id,
#                            invoices=invoices,
#                            hold_types=hold_types, gst_rel=gst_rel, payments=payments, total=total,
#                            current_date=current_date)

@app.route('/contractor_report/<int:contractor_id>')
def contractor_report(contractor_id):
    connection = config.get_db_connection()
    cursor = connection.cursor(dictionary=True, buffered=True)

    try:
        # Fetch contractor details
        cursor.execute("""
            SELECT DISTINCT s.Contractor_Name, st.State_Name, d.District_Name, b.Block_Name,
                   s.Mobile_No, s.GST_Registration_Type, s.GST_No,s.PAN_No,s.Email,s.Address
            FROM subcontractors s
            LEFT JOIN assign_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
            LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
            LEFT JOIN blocks b ON v.Block_Id = b.Block_Id
            LEFT JOIN districts d ON b.District_id = d.District_id
            LEFT JOIN states st ON d.State_Id = st.State_Id
            WHERE s.Contractor_Id = %s
        """, (contractor_id,))
        contInfo = cursor.fetchone()

        # # Fetch distinct hold types present in invoices for the contractor
        # cursor.execute("""
        #     SELECT DISTINCT ht.hold_type_id, ht.hold_type
        #     FROM invoice_subcontractor_hold_join h
        #     JOIN hold_types ht ON h.hold_type_id = ht.hold_type_id
        #     JOIN invoice i ON h.Invoice_Id = i.Invoice_Id
        #     WHERE h.Contractor_Id = %s
        # """, (contractor_id,))
        # hold_types = cursor.fetchall()
        cursor.callproc('GetDistinctHoldTypesInInvoicesByContractor', [contractor_id])

        for result in cursor.stored_results():
            hold_types = result.fetchall()

        # # Fetch all invoices for the contractor, with optional hold information
        # query = """
        #     SELECT DISTINCT i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details,
        #            i.Invoice_Date, i.Invoice_No, i.Basic_Amount, i.Debit_Amount,
        #            i.After_Debit_Amount, i.Amount, i.GST_Amount, i.TDS_Amount, i.SD_Amount,
        #            i.On_Commission, i.Hydro_Testing, i.GST_SD_Amount,
        #            i.Final_Amount, h.hold_amount, ht.hold_type
        #     FROM invoice i
        #     LEFT JOIN villages v ON i.Village_Id = v.Village_Id
        #     LEFT JOIN assign_subcontractors asg ON v.Village_Id = asg.Village_Id
        #     LEFT JOIN subcontractors s ON asg.Contractor_Id = s.Contractor_Id
        #     LEFT JOIN invoice_subcontractor_hold_join h ON i.Invoice_Id = h.Invoice_Id AND h.Contractor_Id = s.Contractor_Id
        #     LEFT JOIN hold_types ht ON h.hold_type_id = ht.hold_type_id
        #     WHERE s.Contractor_Id = %s
        #     ORDER BY i.PMC_No ASC;
        # """
        # cursor.execute(query, (contractor_id,))
        # invoices = cursor.fetchall()
        cursor.callproc('GetInvoicesWithHoldInfoByContractor', [contractor_id])

        for result in cursor.stored_results():
            invoices = result.fetchall()

        # gst_query = """select pmc_no,invoice_no,basic_amount,final_amount from gst_release where pmc_no in
        #                 (select distinct pmc_no from assign_subcontractors where Contractor_Id= %s ) ORDER BY pmc_no ASC """
        # cursor.execute(gst_query, (contractor_id,))
        # gst_rel = cursor.fetchall()
        cursor.callproc('GetGSTReleasesByContractor', [contractor_id])

        for result in cursor.stored_results():
            gst_rel = result.fetchall()

        # pay_query = """select  pmc_no,invoice_no,Payment_Amount,TDS_Payment_Amount,Total_amount,utr from payment where pmc_no in
        #                 (select distinct pmc_no from assign_subcontractors where Contractor_Id=%s ) ORDER BY pmc_no ASC """
        # cursor.execute(pay_query, (contractor_id,))
        # payments = cursor.fetchall()
        cursor.callproc('GetPaymentsByContractor', [contractor_id])

        for result in cursor.stored_results():
            payments = result.fetchall()

        total = {
            "sum_invo_basic_amt": float(sum(row['Basic_Amount'] or 0 for row in invoices)),
            "sum_invo_debit_amt": float(sum(row['Debit_Amount'] or 0 for row in invoices)),
            "sum_invo_after_debit_amt": float(sum(row['After_Debit_Amount'] or 0 for row in invoices)),
            "sum_invo_amt": float(sum(row['Amount'] or 0 for row in invoices)),
            "sum_invo_gst_amt": float(sum(row['GST_Amount'] or 0 for row in invoices)),
            "sum_invo_tds_amt": float(sum(row['TDS_Amount'] or 0 for row in invoices)),
            "sum_invo_ds_amt": float(sum(row['SD_Amount'] or 0 for row in invoices)),
            "sum_invo_on_commission": float(sum(row['On_Commission'] or 0 for row in invoices)),
            "sum_invo_hydro_test": float(sum(row['Hydro_Testing'] or 0 for row in invoices)),
            "sum_invo_gst_sd_amt": float(sum(row['GST_SD_Amount'] or 0 for row in invoices)),
            "sum_invo_final_amt": float(sum(row['Final_Amount'] or 0 for row in invoices)),
            "sum_invo_hold_amt": float(sum(row['hold_amount'] or 0 for row in invoices)),

            "sum_gst_basic_amt": float(sum(row['basic_amount'] or 0 for row in gst_rel)),
            "sum_gst_final_amt": float(sum(row['final_amount'] or 0 for row in gst_rel)),

            "sum_pay_payment_amt": float(sum(row['Payment_Amount'] or 0 for row in payments)),
            "sum_pay_tds_payment_amt": float(sum(row['TDS_Payment_Amount'] or 0 for row in payments)),
            "sum_pay_total_amt": float(sum(row['Total_amount'] or 0 for row in payments))
        }
        current_date = datetime.now().strftime('%Y-%m-%d')

    except Exception as e:
        print(f"Error fetching contractor report: {e}")
        return "An error occurred while fetching contractor report", 500

    finally:
        cursor.close()
        connection.close()

    return render_template('subcontractor_report.html', contInfo=contInfo, contractor_id=contractor_id,
                           invoices=invoices,
                           hold_types=hold_types, gst_rel=gst_rel, payments=payments, total=total,
                           current_date=current_date)


# -----------------------------download contractor report---------------------------------
# @app.route('/download_report/<int:contractor_id>')
# def download_report(contractor_id):
#     connection = config.get_db_connection()
#     cursor = connection.cursor(dictionary=True)
#
#     output_folder = "static/download"
#     output_file = os.path.join(output_folder, f"Contractor_Report_{contractor_id}.xlsx")
#
#     if not os.path.exists(output_folder):
#         os.makedirs(output_folder)
#
#     try:
#         # Fetch Contractor Details
#         # cursor.execute("""
#         #     SELECT DISTINCT s.Contractor_Name, st.State_Name, d.District_Name, b.Block_Name,
#         #                     s.Mobile_No, s.GST_Registration_Type, s.GST_No, s.PAN_No, s.Email, s.Address
#         #     FROM subcontractors s
#         #     LEFT JOIN assign_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
#         #     LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
#         #     LEFT JOIN blocks b ON v.Block_Id = b.Block_Id
#         #     LEFT JOIN districts d ON b.District_id = d.District_id
#         #     LEFT JOIN states st ON d.State_Id = st.State_Id
#         #     WHERE s.Contractor_Id = %s
#         # """, (contractor_id,))
#         # contInfo = cursor.fetchone()
#         cursor.callproc('GetContractorInfoById', [contractor_id])
#         contInfo = next(cursor.stored_results()).fetchone()
#
#         if not contInfo:
#             return "No contractor found", 404
#
#         # Fetch distinct hold types present for the contractor
#         cursor.execute("""
#             SELECT DISTINCT ht.hold_type_id, ht.hold_type
#             FROM invoice_subcontractor_hold_join h
#             JOIN hold_types ht ON h.hold_type_id = ht.hold_type_id
#             WHERE h.Contractor_Id = %s
#         """, (contractor_id,))
#         hold_types = cursor.fetchall()
#         hold_type_map = {ht['hold_type_id']: ht['hold_type'] for ht in hold_types}
#
#         # Fetch Invoices & GST Releases
#         cursor.execute("""
#             SELECT DISTINCT i.Invoice_Id, i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details,
#                    i.Invoice_Date, i.Invoice_No, i.Basic_Amount, i.Debit_Amount,
#                    i.After_Debit_Amount, i.GST_Amount, i.Amount, i.TDS_Amount, i.SD_Amount,
#                    i.On_Commission, i.Hydro_Testing, i.GST_SD_Amount, i.Final_Amount,
#                    g.pmc_no AS gst_pmc_no, g.invoice_no AS gst_invoice_no,
#                    g.basic_amount AS gst_basic_amount, g.final_amount AS gst_final_amount
#             FROM invoice i
#             LEFT JOIN assign_subcontractors asg ON i.PMC_No = asg.PMC_No
#             LEFT JOIN villages v ON i.Village_Id = v.Village_Id
#             LEFT JOIN gst_release g ON i.PMC_No = g.pmc_no AND i.Invoice_No = g.invoice_no
#             WHERE asg.Contractor_Id = %s
#         """, (contractor_id,))
#         invoices = cursor.fetchall()
#
#         # Fetch Hold Amounts separately
#         cursor.execute("""
#             SELECT h.Invoice_Id, ht.hold_type_id, h.hold_amount
#             FROM invoice_subcontractor_hold_join h
#             JOIN hold_types ht ON h.hold_type_id = ht.hold_type_id
#             WHERE h.Contractor_Id = %s
#         """, (contractor_id,))
#         hold_amounts = cursor.fetchall()
#
#         # Create a mapping of invoice_id to hold amounts by type
#         hold_data = {}
#         for h in hold_amounts:
#             hold_data.setdefault(h['Invoice_Id'], {})[h['hold_type_id']] = h['hold_amount']
#
#         # Extract unique PMC numbers for payments
#         pmc_numbers = tuple(set(inv['PMC_No'] for inv in invoices if inv['PMC_No'] is not None))
#
#         # Fetch all Payments for the PMC numbers (including those with null invoice_no)
#         payments_map = {}
#         extra_payments_map = {}  # Now using a map to organize extra payments by PMC
#         if pmc_numbers:
#             # First get payments with invoice_no
#             query = f"""
#                 SELECT pmc_no, invoice_no, Payment_Amount, TDS_Payment_Amount, Total_amount, UTR
#                 FROM payment
#                 WHERE pmc_no IN ({','.join(['%s'] * len(pmc_numbers))})
#                 AND invoice_no IS NOT NULL
#                 ORDER BY pmc_no, invoice_no
#             """
#             cursor.execute(query, pmc_numbers)
#             payments = cursor.fetchall()
#
#             # Organize payments by PMC No & Invoice No
#             for pay in payments:
#                 key = (pay['pmc_no'], pay['invoice_no'])
#                 if key not in payments_map:
#                     payments_map[key] = []
#                 payments_map[key].append(pay)
#
#             # Then get extra payments (null invoice_no) and organize by PMC
#             query = f"""
#                 SELECT pmc_no, invoice_no, Payment_Amount, TDS_Payment_Amount, Total_amount, UTR
#                 FROM payment
#                 WHERE pmc_no IN ({','.join(['%s'] * len(pmc_numbers))})
#                 AND invoice_no IS NULL
#                 ORDER BY pmc_no
#             """
#             cursor.execute(query, pmc_numbers)
#             extra_payments = cursor.fetchall()
#
#             for pay in extra_payments:
#                 if pay['pmc_no'] not in extra_payments_map:
#                     extra_payments_map[pay['pmc_no']] = []
#                 extra_payments_map[pay['pmc_no']].append(pay)
#
#         # Create Excel workbook
#         workbook = openpyxl.Workbook()
#         sheet = workbook.active
#         sheet.title = "Contractor Report"
#
#         # Write Contractor Details
#         sheet.append(["", "", "Laxmi Civil Engineering Services PVT. LTD.", "", ""])
#         sheet.append(
#             ["Contractor Name", contInfo["Contractor_Name"], " ", "GST No", contInfo["GST_No"], " ", "GST Type",
#              contInfo["GST_Registration_Type"]])
#         sheet.append(
#             ["State", contInfo["State_Name"], " ", "PAN No", contInfo["PAN_No"], " ", "Address", contInfo["Address"]])
#         sheet.append(["District", contInfo["District_Name"], " ", "Mobile No", contInfo["Mobile_No"]])
#         sheet.append(["Block", contInfo["Block_Name"], " ", "Email", contInfo["Email"]])
#         sheet.append([])
#
#         # Table Headers - include all hold types as separate columns
#         base_headers = ["PMC No", "Village", "Work Type", "Invoice Details", "Invoice Date", "Invoice No",
#                         "Basic Amount", "Debit", "After Debit Amount", "GST (18%)", "Amount", "TDS (1%)",
#                         "SD (5%)", "On Commission", "Hydro Testing", "GST SD Amount"]
#
#         hold_headers = [ht['hold_type'] for ht in hold_types]
#
#         payment_headers = ["Final Amount", "Payment Amount", "TDS Payment", "Total Paid", "UTR"]
#
#         sheet.append(base_headers + hold_headers + payment_headers)
#
#         seen_invoices = set()
#         seen_gst_notes = set()
#         processed_payments = set()  # Track which payments we've processed
#
#         # Process invoices grouped by PMC No
#         pmc_groups = {}
#         for inv in invoices:
#             pmc_no = inv["PMC_No"]
#             if pmc_no not in pmc_groups:
#                 pmc_groups[pmc_no] = []
#             pmc_groups[pmc_no].append(inv)
#
#         # Process each PMC group separately
#         for pmc_no, pmc_invoices in pmc_groups.items():
#             # Process all invoices for this PMC first
#             for inv in pmc_invoices:
#                 invoice_no = inv["Invoice_No"]
#                 payments = payments_map.get((pmc_no, invoice_no), [])
#
#                 # Process invoice row with first payment (if exists)
#                 if (pmc_no, invoice_no) not in seen_invoices:
#                     seen_invoices.add((pmc_no, invoice_no))
#                     first_payment = payments[0] if len(payments) > 0 else None
#
#                     # Base invoice data
#                     row = [
#                         pmc_no, inv["Village_Name"], inv["Work_Type"], inv["Invoice_Details"],
#                         inv["Invoice_Date"], invoice_no, inv["Basic_Amount"], inv["Debit_Amount"],
#                         inv["After_Debit_Amount"], inv["GST_Amount"], inv["Amount"], inv["TDS_Amount"],
#                         inv["SD_Amount"], inv["On_Commission"], inv["Hydro_Testing"], inv["GST_SD_Amount"]
#                     ]
#
#                     # Add hold amounts for each hold type
#                     invoice_holds = hold_data.get(inv["Invoice_Id"], {})
#                     for ht_id in hold_type_map.keys():
#                         row.append(invoice_holds.get(ht_id, ""))
#
#                     # Add payment information
#                     row += [
#                         inv["Final_Amount"],
#                         first_payment["Payment_Amount"] if first_payment else "",
#                         first_payment["TDS_Payment_Amount"] if first_payment else "",
#                         first_payment["Total_amount"] if first_payment else "",
#                         first_payment["UTR"] if first_payment else ""
#                     ]
#
#                     sheet.append(row)
#
#                     if first_payment:
#                         payment_id = f"{pmc_no}-{invoice_no}-{first_payment['Payment_Amount']}-{first_payment.get('UTR', '')}"
#                         processed_payments.add(payment_id)
#
#                 # Process GST release if exists (only if we have a matching GST record)
#                 if inv["gst_pmc_no"] and (inv["gst_pmc_no"], inv["gst_invoice_no"]) not in seen_gst_notes:
#                     seen_gst_notes.add((inv["gst_pmc_no"], inv["gst_invoice_no"]))
#
#                     # Find the payment that matches this GST release
#                     gst_payment = None
#                     for payment in payments[1:]:  # Skip first payment (already used for invoice)
#                         if payment['invoice_no'] == inv["gst_invoice_no"]:
#                             gst_payment = payment
#                             break
#
#                     # GST release row
#                     row = [
#                         pmc_no, "", "", "GST Release Note", "", inv["gst_invoice_no"],
#                         inv["gst_basic_amount"], "", "", "", "", "", "", "", "", ""  # Empty GST SD Amount
#                     ]
#
#                     # Empty holds for GST release
#                     row += ["" for _ in hold_headers]
#
#                     # Add payment information
#                     row += [
#                         inv["gst_final_amount"],
#                         gst_payment["Payment_Amount"] if gst_payment else "",
#                         gst_payment["TDS_Payment_Amount"] if gst_payment else "",
#                         gst_payment["Total_amount"] if gst_payment else "",
#                         gst_payment["UTR"] if gst_payment else ""
#                     ]
#
#                     sheet.append(row)
#
#                     if gst_payment:
#                         payment_id = f"{pmc_no}-{inv['gst_invoice_no']}-{gst_payment['Payment_Amount']}-{gst_payment.get('UTR', '')}"
#                         processed_payments.add(payment_id)
#
#                 # Process remaining payments as extra payments
#                 for payment in payments[1:]:
#                     payment_id = f"{pmc_no}-{payment['invoice_no']}-{payment['Payment_Amount']}-{payment.get('UTR', '')}"
#                     if payment_id not in processed_payments:
#                         row = [
#                             pmc_no, "", "", "", "", payment['invoice_no'],
#                             "", "", "", "", "", "", "", "", "", ""  # Empty GST SD Amount
#                         ]
#
#                         # Empty holds for extra payments
#                         row += ["" for _ in hold_headers]
#
#                         # Add payment information
#                         row += [
#                             "",
#                             payment["Payment_Amount"],
#                             payment["TDS_Payment_Amount"],
#                             payment["Total_amount"],
#                             payment["UTR"]
#                         ]
#
#                         sheet.append(row)
#                         processed_payments.add(payment_id)
#
#             # Process extra payments for this PMC (null invoice_no) immediately after the PMC's invoices
#             if pmc_no in extra_payments_map:
#                 for payment in extra_payments_map[pmc_no]:
#                     payment_id = f"{pmc_no}-null-{payment['Payment_Amount']}-{payment.get('UTR', '')}"
#                     if payment_id not in processed_payments:
#                         row = [
#                             pmc_no, "", "", "", "", "",
#                             "", "", "", "", "", "", "", "", "", ""  # Empty GST SD Amount
#                         ]
#
#                         # Empty holds for null invoice payments
#                         row += ["" for _ in hold_headers]
#
#                         # Add payment information
#                         row += [
#                             "",
#                             payment["Payment_Amount"],
#                             payment["TDS_Payment_Amount"],
#                             payment["Total_amount"],
#                             payment["UTR"]
#                         ]
#
#                         sheet.append(row)
#                         processed_payments.add(payment_id)
#
#         workbook.save(output_file)
#         workbook.close()
#     finally:
#         cursor.close()
#         connection.close()
#     from openpyxl.styles import Font
#
#     # Initialize totals
#     total_basic_amount = 0
#     total_tds_amount = 0
#     total_sd_amount = 0
#     total_on_commission = 0
#     total_hold_amount = 0
#     total_final_amount = 0
#     total_total_amount = 0  # Sum of all Total Amounts from payment data
#
#     # Iterate through rows to calculate totals
#     for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
#         try:
#             total_basic_amount += float(row[6] or 0)  # Basic_Amount
#             total_tds_amount += float(row[11] or 0)  # TDS_Amount
#             total_sd_amount += float(row[12] or 0)  # SD_Amount
#             total_on_commission += float(row[13] or 0)  # On_Commission
#             total_final_amount += float(row[-5] or 0)  # Final_Amount
#             total_total_amount += float(row[-2] or 0)  # Total_Amount (from payments)
#
#             # Sum of hold amounts (dynamically calculated based on hold_headers)
#             hold_start_col = len(base_headers)  # First column where hold types start
#             hold_end_col = hold_start_col + len(hold_headers)  # Last column where hold types end
#             total_hold_amount += sum(float(row[i] or 0) for i in range(hold_start_col, hold_end_col))
#
#         except ValueError:
#             continue  # Skip if non-numeric values cause errors
#
#     # Append the totals row at the bottom of the sheet
#     totals_row = [
#         "TOTAL", "", "", "", "", "",  # Empty values for non-numeric columns
#         total_basic_amount, "", "", "", "", total_tds_amount, total_sd_amount,
#         total_on_commission, "", "",  # Empty GST SD Amount
#     ]
#
#     # Add empty placeholders for hold types
#     totals_row += [total_hold_amount] + [""] * (len(hold_headers) - 1)
#
#     # Add totals for Final Amount and Total Amount
#     totals_row += [
#         total_final_amount, "", "", total_total_amount, ""  # UTR column remains empty
#     ]
#
#     # Append totals row to sheet
#     sheet.append([])
#     sheet.append(totals_row)
#
#     # Make the totals row bold
#     for cell in sheet[sheet.max_row]:
#         cell.font = Font(bold=True)
#
#     sheet.append([])
#     sheet.append([])
#     sheet.append([])
#     sheet.append([])
#     sheet.append(["This is Generated by LCEPL Application.."])
#
#     # Save workbook after modifications
#     workbook.save(output_file)
#     workbook.close()
#
#     return send_from_directory(output_folder, f"Contractor_Report_{contractor_id}.xlsx", as_attachment=True)
from openpyxl import Workbook

# # Download report by contractor id
# # Download report by contractor id
from flask import send_from_directory
import os
import openpyxl
from openpyxl.styles import Font
import config

from flask import send_from_directory
import os
import openpyxl
from openpyxl.utils import get_column_letter
import config

# -----------------------PMC Report---------------------------
@app.route('/download_report/<int:contractor_id>')
def download_report(contractor_id):
    connection = config.get_db_connection()
    cursor = connection.cursor(dictionary=True)

    output_folder = "static/download"
    output_file = os.path.join(output_folder, f"Contractor_Report_{contractor_id}.xlsx")

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    try:
        # Fetch Contractor Details
        cursor.execute("""
            SELECT DISTINCT s.Contractor_Name, st.State_Name, d.District_Name, b.Block_Name,
                            s.Mobile_No, s.GST_Registration_Type, s.GST_No, s.PAN_No, s.Email, s.Address
            FROM subcontractors s
            LEFT JOIN assign_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
            LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
            LEFT JOIN blocks b ON v.Block_Id = b.Block_Id
            LEFT JOIN districts d ON b.District_id = d.District_id
            LEFT JOIN states st ON d.State_Id = st.State_Id
            WHERE s.Contractor_Id = %s
        """, (contractor_id,))
        contInfo = cursor.fetchone()

        if not contInfo:
            return "No contractor found", 404

        # # Fetch distinct hold types present for the contractor
        # cursor.execute("""
        #     SELECT DISTINCT ht.hold_type_id, ht.hold_type
        #     FROM invoice_subcontractor_hold_join h
        #     JOIN hold_types ht ON h.hold_type_id = ht.hold_type_id
        #     WHERE h.Contractor_Id = %s
        # """, (contractor_id,))
        # hold_types = cursor.fetchall()
        cursor.callproc('GetDistinctHoldTypesByContractor', [contractor_id])

        for result in cursor.stored_results():
            hold_types = result.fetchall()

        hold_type_map = {ht['hold_type_id']: ht['hold_type'] for ht in hold_types}

        # # Fetch Invoices & GST Releases
        # cursor.execute("""
        #     SELECT DISTINCT i.Invoice_Id, i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details,
        #            i.Invoice_Date, i.Invoice_No, i.Basic_Amount, i.Debit_Amount,
        #            i.After_Debit_Amount, i.GST_Amount, i.Amount, i.TDS_Amount, i.SD_Amount,
        #            i.On_Commission, i.Hydro_Testing, i.GST_SD_Amount, i.Final_Amount,
        #            g.pmc_no AS gst_pmc_no, g.invoice_no AS gst_invoice_no,
        #            g.basic_amount AS gst_basic_amount, g.final_amount AS gst_final_amount
        #     FROM invoice i
        #     LEFT JOIN assign_subcontractors asg ON i.PMC_No = asg.PMC_No
        #     LEFT JOIN villages v ON i.Village_Id = v.Village_Id
        #     LEFT JOIN gst_release g ON i.PMC_No = g.pmc_no AND i.Invoice_No = g.invoice_no
        #     WHERE asg.Contractor_Id = %s
        # """, (contractor_id,))
        # invoices = cursor.fetchall()
        cursor.callproc('GetInvoicesAndGSTReleasesByContractor', [contractor_id])

        for result in cursor.stored_results():
            invoices = result.fetchall()

        # # Fetch Hold Amounts separately
        # cursor.execute("""
        #     SELECT h.Invoice_Id, ht.hold_type_id, h.hold_amount
        #     FROM invoice_subcontractor_hold_join h
        #     JOIN hold_types ht ON h.hold_type_id = ht.hold_type_id
        #     WHERE h.Contractor_Id = %s
        # """, (contractor_id,))
        # hold_amounts = cursor.fetchall()
        cursor.callproc('GetHoldAmountsByContractors', [contractor_id])

        for result in cursor.stored_results():
            hold_amounts = result.fetchall()

        # Create a mapping of invoice_id to hold amounts by type
        hold_data = {}
        for h in hold_amounts:
            hold_data.setdefault(h['Invoice_Id'], {})[h['hold_type_id']] = h['hold_amount']

        # Extract unique PMC numbers for payments
        pmc_numbers = tuple(set(inv['PMC_No'] for inv in invoices if inv['PMC_No'] is not None))

        # Fetch all Payments for the PMC numbers (including those with null invoice_no)
        payments_map = {}
        extra_payments_map = {}  # Now using a map to organize extra payments by PMC
        if pmc_numbers:
            # First get payments with invoice_no
            query = f"""
                SELECT pmc_no, invoice_no, Payment_Amount, TDS_Payment_Amount, Total_amount, UTR
                FROM payment
                WHERE pmc_no IN ({','.join(['%s'] * len(pmc_numbers))})
                AND invoice_no IS NOT NULL
                ORDER BY pmc_no, invoice_no
            """
            cursor.execute(query, pmc_numbers)
            payments = cursor.fetchall()

            # Organize payments by PMC No & Invoice No
            for pay in payments:
                key = (pay['pmc_no'], pay['invoice_no'])
                if key not in payments_map:
                    payments_map[key] = []
                payments_map[key].append(pay)

            # Then get extra payments (null invoice_no) and organize by PMC
            query = f"""
                SELECT pmc_no, invoice_no, Payment_Amount, TDS_Payment_Amount, Total_amount, UTR
                FROM payment
                WHERE pmc_no IN ({','.join(['%s'] * len(pmc_numbers))})
                AND invoice_no IS NULL
                ORDER BY pmc_no
            """
            cursor.execute(query, pmc_numbers)
            extra_payments = cursor.fetchall()

            for pay in extra_payments:
                if pay['pmc_no'] not in extra_payments_map:
                    extra_payments_map[pay['pmc_no']] = []
                extra_payments_map[pay['pmc_no']].append(pay)

        # Create Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Contractor Report"

        # Write Contractor Details
        sheet.append(["Contractor Name", contInfo["Contractor_Name"]])
        sheet.append(["State", contInfo["State_Name"]])
        sheet.append(["District", contInfo["District_Name"]])
        sheet.append(["Block", contInfo["Block_Name"]])
        sheet.append(["Mobile No", contInfo["Mobile_No"]])
        sheet.append(["GST Type", contInfo["GST_Registration_Type"]])
        sheet.append(["GST No", contInfo["GST_No"]])
        sheet.append(["PAN No", contInfo["PAN_No"]])
        sheet.append(["Email", contInfo["Email"]])
        sheet.append(["Address", contInfo["Address"]])
        sheet.append([])

        # Table Headers - include all hold types as separate columns
        base_headers = ["PMC No", "Village", "Work Type", "Invoice Details", "Invoice Date", "Invoice No",
                        "Basic Amount", "Debit", "After Debit Amount", "GST (18%)", "Amount", "TDS (1%)",
                        "SD (5%)", "On Commission", "Hydro Testing", "GST SD Amount"]

        hold_headers = [ht['hold_type'] for ht in hold_types]

        payment_headers = ["Final Amount", "Payment Amount", "TDS Payment", "Total Paid", "UTR"]

        sheet.append(base_headers + hold_headers + payment_headers)

        seen_invoices = set()
        seen_gst_notes = set()
        processed_payments = set()  # Track which payments we've processed

        # Process invoices grouped by PMC No
        pmc_groups = {}
        for inv in invoices:
            pmc_no = inv["PMC_No"]
            if pmc_no not in pmc_groups:
                pmc_groups[pmc_no] = []
            pmc_groups[pmc_no].append(inv)

        # Process each PMC group separately
        for pmc_no, pmc_invoices in pmc_groups.items():
            # Process all invoices for this PMC first
            for inv in pmc_invoices:
                invoice_no = inv["Invoice_No"]
                payments = payments_map.get((pmc_no, invoice_no), [])

                # Process invoice row with first payment (if exists)
                if (pmc_no, invoice_no) not in seen_invoices:
                    seen_invoices.add((pmc_no, invoice_no))
                    first_payment = payments[0] if len(payments) > 0 else None

                    # Base invoice data
                    row = [
                        pmc_no, inv["Village_Name"], inv["Work_Type"], inv["Invoice_Details"],
                        inv["Invoice_Date"], invoice_no, inv["Basic_Amount"], inv["Debit_Amount"],
                        inv["After_Debit_Amount"], inv["GST_Amount"], inv["Amount"], inv["TDS_Amount"],
                        inv["SD_Amount"], inv["On_Commission"], inv["Hydro_Testing"], inv["GST_SD_Amount"]
                    ]

                    # Add hold amounts for each hold type
                    invoice_holds = hold_data.get(inv["Invoice_Id"], {})
                    for ht_id in hold_type_map.keys():
                        row.append(invoice_holds.get(ht_id, ""))

                    # Add payment information
                    row += [
                        inv["Final_Amount"],
                        first_payment["Payment_Amount"] if first_payment else "",
                        first_payment["TDS_Payment_Amount"] if first_payment else "",
                        first_payment["Total_amount"] if first_payment else "",
                        first_payment["UTR"] if first_payment else ""
                    ]

                    sheet.append(row)

                    if first_payment:
                        payment_id = f"{pmc_no}-{invoice_no}-{first_payment['Payment_Amount']}-{first_payment.get('UTR', '')}"
                        processed_payments.add(payment_id)

                # Process GST release if exists (only if we have a matching GST record)
                if inv["gst_pmc_no"] and (inv["gst_pmc_no"], inv["gst_invoice_no"]) not in seen_gst_notes:
                    seen_gst_notes.add((inv["gst_pmc_no"], inv["gst_invoice_no"]))

                    # Find the payment that matches this GST release
                    gst_payment = None
                    for payment in payments[1:]:  # Skip first payment (already used for invoice)
                        if payment['invoice_no'] == inv["gst_invoice_no"]:
                            gst_payment = payment
                            break

                    # GST release row
                    row = [
                        pmc_no, "", "", "GST Release Note", "", inv["gst_invoice_no"],
                        inv["gst_basic_amount"], "", "", "", "", "", "", "", "", ""  # Empty GST SD Amount
                    ]

                    # Empty holds for GST release
                    row += ["" for _ in hold_headers]

                    # Add payment information
                    row += [
                        inv["gst_final_amount"],
                        gst_payment["Payment_Amount"] if gst_payment else "",
                        gst_payment["TDS_Payment_Amount"] if gst_payment else "",
                        gst_payment["Total_amount"] if gst_payment else "",
                        gst_payment["UTR"] if gst_payment else ""
                    ]

                    sheet.append(row)

                    if gst_payment:
                        payment_id = f"{pmc_no}-{inv['gst_invoice_no']}-{gst_payment['Payment_Amount']}-{gst_payment.get('UTR', '')}"
                        processed_payments.add(payment_id)

                # Process remaining payments as extra payments
                for payment in payments[1:]:
                    payment_id = f"{pmc_no}-{payment['invoice_no']}-{payment['Payment_Amount']}-{payment.get('UTR', '')}"
                    if payment_id not in processed_payments:
                        row = [
                            pmc_no, "", "", "", "", payment['invoice_no'],
                            "", "", "", "", "", "", "", "", "", ""  # Empty GST SD Amount
                        ]

                        # Empty holds for extra payments
                        row += ["" for _ in hold_headers]

                        # Add payment information
                        row += [
                            "",
                            payment["Payment_Amount"],
                            payment["TDS_Payment_Amount"],
                            payment["Total_amount"],
                            payment["UTR"]
                        ]

                        sheet.append(row)
                        processed_payments.add(payment_id)

            # Process extra payments for this PMC (null invoice_no) immediately after the PMC's invoices
            if pmc_no in extra_payments_map:
                for payment in extra_payments_map[pmc_no]:
                    payment_id = f"{pmc_no}-null-{payment['Payment_Amount']}-{payment.get('UTR', '')}"
                    if payment_id not in processed_payments:
                        row = [
                            pmc_no, "", "", "", "", "",
                            "", "", "", "", "", "", "", "", "", ""  # Empty GST SD Amount
                        ]

                        # Empty holds for null invoice payments
                        row += ["" for _ in hold_headers]

                        # Add payment information
                        row += [
                            "",
                            payment["Payment_Amount"],
                            payment["TDS_Payment_Amount"],
                            payment["Total_amount"],
                            payment["UTR"]
                        ]


                        sheet.append(row)
                        processed_payments.add(payment_id)

        workbook.save(output_file)
        workbook.close()
    finally:
        cursor.close()
        connection.close()
    from openpyxl.styles import Font

    # Initialize totals
    total_basic_amount = 0
    total_tds_amount = 0
    total_sd_amount = 0
    total_on_commission = 0
    total_hold_amount = 0
    total_final_amount = 0
    total_total_amount = 0  # Sum of all Total Amounts from payment data

    # Iterate through rows to calculate totals
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        try:
            total_basic_amount += float(row[6] or 0)  # Basic_Amount
            total_tds_amount += float(row[11] or 0)  # TDS_Amount
            total_sd_amount += float(row[12] or 0)  # SD_Amount
            total_on_commission += float(row[13] or 0)  # On_Commission
            total_final_amount += float(row[-5] or 0)  # Final_Amount
            total_total_amount += float(row[-2] or 0)  # Total_Amount (from payments)

            # Sum of hold amounts (dynamically calculated based on hold_headers)
            hold_start_col = len(base_headers)  # First column where hold types start
            hold_end_col = hold_start_col + len(hold_headers)  # Last column where hold types end
            total_hold_amount += sum(float(row[i] or 0) for i in range(hold_start_col, hold_end_col))

        except ValueError:
            continue  # Skip if non-numeric values cause errors

    # Append the totals row at the bottom of the sheet
    totals_row = [
        "TOTAL", "", "", "", "", "",  # Empty values for non-numeric columns
        total_basic_amount, "", "", "", "", total_tds_amount, total_sd_amount,
        total_on_commission, "", "",  # Empty GST SD Amount
    ]

    # Add empty placeholders for hold types
    totals_row += [total_hold_amount] + [""] * (len(hold_headers) - 1)

    # Add totals for Final Amount and Total Amount
    totals_row += [
        total_final_amount, "", "", total_total_amount, ""  # UTR column remains empty
    ]

    # Append totals row to sheet
    sheet.append([])
    sheet.append(totals_row)

    # Make the totals row bold
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)

    # Save workbook after modifications
    workbook.save(output_file)
    workbook.close()

    return send_from_directory(output_folder, f"Contractor_Report_{contractor_id}.xlsx", as_attachment=True)

# show report by pmc no
@app.route('/pmc_report/<pmc_no>')
def pmc_report(pmc_no):
    connection = config.get_db_connection()
    cursor = connection.cursor(dictionary=True, buffered=True)

    try:
        # Fetch PMC details and contractor information
        # cursor.execute("""
        #     SELECT DISTINCT a.PMC_No, a.Village_Id, v.Village_Name, b.Block_Name,
        #            d.District_Name, s.State_Name, sc.Contractor_Id, sc.Contractor_Name,
        #            sc.Address, sc.Mobile_No, sc.PAN_No, sc.Email, sc.Gender,
        #            sc.GST_Registration_Type, sc.GST_No
        #     FROM assign_subcontractors a
        #     INNER JOIN villages v ON a.Village_Id = v.Village_Id
        #     INNER JOIN blocks b ON v.Block_Id = b.Block_Id
        #     INNER JOIN districts d ON b.District_id = d.District_id
        #     INNER JOIN states s ON d.State_Id = s.State_Id
        #     INNER JOIN subcontractors sc ON a.Contractor_Id = sc.Contractor_Id
        #     WHERE a.pmc_no = %s
        # """, (pmc_no,))
        # pmc_info = cursor.fetchone()

        cursor.callproc("GetContractorInfoByPmcNo", (pmc_no,))
        for pmc in cursor.stored_results():
            pmc_info = pmc.fetchone()

        if not pmc_info:
            return "No PMC found with this number", 404

        # Fetch distinct hold types present in invoices for this PMC
        # cursor.execute("""
        #     SELECT DISTINCT ht.hold_type_id, ht.hold_type
        #     FROM invoice_subcontractor_hold_join h
        #     JOIN hold_types ht ON h.hold_type_id = ht.hold_type_id
        #     JOIN invoice i ON h.Invoice_Id = i.Invoice_Id
        #     JOIN assign_subcontractors a ON i.PMC_No = a.PMC_No
        #     WHERE a.PMC_No = %s AND a.Contractor_Id = %s
        # """, (pmc_no, pmc_info["Contractor_Id"]))
        # hold_types = cursor.fetchall()
        cursor.callproc("Get_pmc_hold_types", (pmc_no, pmc_info["Contractor_Id"]))
        for hold in cursor.stored_results():
            hold_types = hold.fetchall()

        # Extract hold type IDs
        hold_type_ids = [ht['hold_type_id'] for ht in hold_types]

        # Initialize variables
        invoices = []
        gst_rel = []
        payments = []
        hold_amount_total = 0
        total_invo_final = 0
        total_gst_basic = 0
        total_gst_final = 0
        total_pay_amount = 0
        total_pay_total = 0

        if not hold_type_ids:
            # # Query without hold types
            # query = """
            #     SELECT DISTINCT i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details,
            #            i.Invoice_Date, i.Invoice_No, i.Basic_Amount, i.Debit_Amount,
            #            i.After_Debit_Amount, i.Amount, i.GST_Amount, i.TDS_Amount, i.SD_Amount,
            #            i.On_Commission, i.Hydro_Testing, i.GST_SD_Amount, i.Final_Amount
            #     FROM invoice i
            #     LEFT JOIN villages v ON i.Village_Id = v.Village_Id
            #     LEFT JOIN assign_subcontractors a ON i.PMC_No = a.PMC_No
            #     WHERE a.PMC_No = %s AND a.Contractor_Id = %s
            #     ORDER BY i.Invoice_Date, i.Invoice_No
            # """
            # cursor.execute(query, (pmc_no, pmc_info["Contractor_Id"]))
            # invoices = cursor.fetchall()
            cursor.callproc('GetInvoicesByPMCAndContractor', [pmc_no, pmc_info["Contractor_Id"]])

            # Fetch results
            for result in cursor.stored_results():
                invoices = result.fetchall()

        else:
            # Query with hold types
            query = """
                SELECT DISTINCT i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details,
                       i.Invoice_Date, i.Invoice_No, i.Basic_Amount, i.Debit_Amount,
                       i.After_Debit_Amount, i.Amount, i.GST_Amount, i.TDS_Amount, i.SD_Amount,
                       i.On_Commission, i.Hydro_Testing, i.GST_SD_Amount,
                       i.Final_Amount, h.hold_amount, ht.hold_type
                FROM invoice i
                LEFT JOIN villages v ON i.Village_Id = v.Village_Id
                LEFT JOIN assign_subcontractors a ON i.PMC_No = a.PMC_No
                LEFT JOIN invoice_subcontractor_hold_join h ON i.Invoice_Id = h.Invoice_Id
                LEFT JOIN hold_types ht ON h.hold_type_id = ht.hold_type_id
                WHERE a.PMC_No = %s AND a.Contractor_Id = %s
                AND (ht.hold_type_id IS NULL OR ht.hold_type_id IN ({0}))
                ORDER BY i.Invoice_Date, i.Invoice_No
            """.format(','.join(['%s'] * len(hold_type_ids)))

            params = [pmc_no, pmc_info["Contractor_Id"]] + hold_type_ids
            cursor.execute(query, params)
            invoices = cursor.fetchall()
            hold_amount_total = sum(row['hold_amount'] or 0 for row in invoices if row['hold_amount'] is not None)

        # Calculate invoice totals
        total_invo_final = float(sum(row['Final_Amount'] or 0 for row in invoices))

        # # GST Release query
        # gst_query = """
        #     SELECT pmc_no, invoice_no, basic_amount, final_amount
        #     FROM gst_release
        #     WHERE pmc_no = %s
        #     ORDER BY invoice_no ASC
        # """
        # cursor.execute(gst_query, (pmc_no,))

        # gst_rel = cursor.fetchall()
        cursor.callproc('GetGSTReleaseByPMC', [pmc_no])

        # Fetch results
        for result in cursor.stored_results():
            gst_rel = result.fetchall()

        total_gst_basic = float(sum(row['basic_amount'] or 0 for row in gst_rel))
        total_gst_final = float(sum(row['final_amount'] or 0 for row in gst_rel))

        # # Payment query
        # pay_query = """
        #     SELECT pmc_no, invoice_no, Payment_Amount, TDS_Payment_Amount, Total_amount, utr
        #     FROM payment
        #     WHERE pmc_no = %s
        #     ORDER BY invoice_no ASC
        # """
        # cursor.execute(pay_query, (pmc_no,))
        # payments = cursor.fetchall()
        cursor.callproc('GetPaymentByPMC', [pmc_no])

        for result in cursor.stored_results():
            payments = result.fetchall()

        total_pay_amount = float(sum(row['Payment_Amount'] or 0 for row in payments))
        total_pay_total = float(sum(row['Total_amount'] or 0 for row in payments))

        # Prepare totals dictionary
        totals = {
            "sum_invo_basic_amt": float(sum(row['Basic_Amount'] or 0 for row in invoices)),
            "sum_invo_debit_amt": float(sum(row['Debit_Amount'] or 0 for row in invoices)),
            "sum_invo_after_debit_amt": float(sum(row['After_Debit_Amount'] or 0 for row in invoices)),
            "sum_invo_amt": float(sum(row['Amount'] or 0 for row in invoices)),
            "sum_invo_gst_amt": float(sum(row['GST_Amount'] or 0 for row in invoices)),
            "sum_invo_tds_amt": float(sum(row['TDS_Amount'] or 0 for row in invoices)),
            "sum_invo_ds_amt": float(sum(row['SD_Amount'] or 0 for row in invoices)),
            "sum_invo_on_commission": float(sum(row['On_Commission'] or 0 for row in invoices)),
            "sum_invo_hydro_test": float(sum(row['Hydro_Testing'] or 0 for row in invoices)),
            "sum_invo_gst_sd_amt": float(sum(row['GST_SD_Amount'] or 0 for row in invoices)),
            "sum_invo_final_amt": total_invo_final,
            "sum_invo_hold_amt": hold_amount_total,
            "sum_gst_basic_amt": total_gst_basic,
            "sum_gst_final_amt": total_gst_final,
            "sum_pay_payment_amt": total_pay_amount,
            "sum_pay_tds_payment_amt": float(sum(row['TDS_Payment_Amount'] or 0 for row in payments)),
            "sum_pay_total_amt": total_pay_total
        }

    except Exception as e:
        print(f"Error fetching PMC report: {e}")
        return "An error occurred while fetching PMC report", 500

    finally:
        cursor.close()
        connection.close()

    return render_template('pmc_report.html',
                           info=pmc_info,
                           invoices=invoices,
                           hold_types=hold_types,
                           gst_rel=gst_rel,
                           payments=payments,
                           total=totals)


# Download report by PMC No
@app.route('/download_pmc_report/<pmc_no>')
def download_pmc_report(pmc_no):
    connection = config.get_db_connection()
    output_folder = "static/download"
    output_file = os.path.join(output_folder, f"PMC_Report_{pmc_no}.xlsx")

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    cursor = connection.cursor(dictionary=True)

    try:
        # # Fetch Contractor Details using PMC No
        # cursor.execute("""
        #     SELECT DISTINCT s.Contractor_Id, s.Contractor_Name, st.State_Name, d.District_Name, b.Block_Name,
        #                     s.Mobile_No, s.GST_Registration_Type, s.GST_No, s.PAN_No, s.Email, s.Address
        #     FROM subcontractors s
        #     LEFT JOIN assign_subcontractors asg ON s.Contractor_Id = asg.Contractor_Id
        #     LEFT JOIN villages v ON asg.Village_Id = v.Village_Id
        #     LEFT JOIN blocks b ON v.Block_Id = b.Block_Id
        #     LEFT JOIN districts d ON b.District_id = d.District_id
        #     LEFT JOIN states st ON d.State_Id = st.State_Id
        #     WHERE asg.PMC_No = %s
        # """, (pmc_no,))
        # contractor_info = cursor.fetchone()
        cursor.callproc('GetContractorDetailsByPMC', [pmc_no])

        # Now fetch the result:
        for result in cursor.stored_results():
            contractor_info = result.fetchone()

        if not contractor_info:
            return "No contractor found for this PMC No", 404

        # # Fetch distinct hold types present for the contractor
        # cursor.execute("""
        #     SELECT DISTINCT ht.hold_type_id, ht.hold_type
        #     FROM invoice_subcontractor_hold_join h
        #     JOIN hold_types ht ON h.hold_type_id = ht.hold_type_id
        #     WHERE h.Contractor_Id = %s
        # """, (contractor_info["Contractor_Id"],))
        # hold_types = cursor.fetchall()
        cursor.callproc('GetHoldTypesByContractor', [contractor_info["Contractor_Id"]])

        for result in cursor.stored_results():
            hold_types = result.fetchall()

        hold_type_map = {ht['hold_type_id']: ht['hold_type'] for ht in hold_types}

        # # Fetch Invoices & GST Releases
        # cursor.execute("""
        #     SELECT DISTINCT i.Invoice_Id, i.PMC_No, v.Village_Name, i.Work_Type, i.Invoice_Details,
        #            i.Invoice_Date, i.Invoice_No, i.Basic_Amount, i.Debit_Amount,
        #            i.After_Debit_Amount, i.GST_Amount, i.Amount, i.TDS_Amount, i.SD_Amount,
        #            i.On_Commission, i.Hydro_Testing, i.GST_SD_Amount, i.Final_Amount,
        #            g.pmc_no AS gst_pmc_no, g.invoice_no AS gst_invoice_no,
        #            g.basic_amount AS gst_basic_amount, g.final_amount AS gst_final_amount
        #     FROM invoice i
        #     LEFT JOIN assign_subcontractors asg ON i.PMC_No = asg.PMC_No
        #     LEFT JOIN villages v ON i.Village_Id = v.Village_Id
        #     LEFT JOIN gst_release g ON i.PMC_No = g.pmc_no AND i.Invoice_No = g.invoice_no
        #     WHERE asg.PMC_No = %s
        #     ORDER BY i.Invoice_Date, i.Invoice_No
        # """, (pmc_no,))
        # invoices = cursor.fetchall()
        cursor.callproc('GetInvoicesAndGSTReleasesByPMC', [pmc_no])

        for result in cursor.stored_results():
            invoices = result.fetchall()

        # # Fetch Hold Amounts separately
        # cursor.execute("""
        #     SELECT h.Invoice_Id, ht.hold_type_id, h.hold_amount
        #     FROM invoice_subcontractor_hold_join h
        #     JOIN hold_types ht ON h.hold_type_id = ht.hold_type_id
        #     WHERE h.Contractor_Id = %s
        # """, (contractor_info["Contractor_Id"],))
        # hold_amounts = cursor.fetchall()
        cursor.callproc('GetHoldAmountsByContractor', [contractor_info["Contractor_Id"]])

        for result in cursor.stored_results():
            hold_amounts = result.fetchall()

        # Create a mapping of invoice_id to hold amounts by type
        hold_data = {}
        for h in hold_amounts:
            hold_data.setdefault(h['Invoice_Id'], {})[h['hold_type_id']] = h['hold_amount']

        # # Fetch all Payments for the PMC number
        # cursor.execute("""
        #     SELECT pmc_no, invoice_no, Payment_Amount, TDS_Payment_Amount, Total_amount, UTR
        #     FROM payment
        #     WHERE pmc_no = %s
        #     ORDER BY invoice_no
        # """, (pmc_no,))
        # all_payments = cursor.fetchall()
        cursor.callproc('GetAllPaymentsByPMC', [pmc_no])

        for result in cursor.stored_results():
            all_payments = result.fetchall()

        # Organize payments by Invoice No (both regular and GST release notes)
        payments_map = {}
        extra_payments = []
        for pay in all_payments:
            if pay['invoice_no']:
                key = pay['invoice_no']
                if key not in payments_map:
                    payments_map[key] = []
                payments_map[key].append(pay)
            else:
                extra_payments.append(pay)

        # Create Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "PMC Report"

        # Write Contractor Details
        sheet.append(["", "", "Laxmi Civil Engineering Services PVT. LTD.", "", ""])
        sheet.append(
            ["Contractor Name", contractor_info["Contractor_Name"], " ", "GST No", contractor_info["GST_No"], " ",
             "GST Type", contractor_info["GST_Registration_Type"]])
        sheet.append(["State", contractor_info["State_Name"], " ", "PAN No", contractor_info["PAN_No"], " ", "Address",
                      contractor_info["Address"]])
        sheet.append(["District", contractor_info["District_Name"], " ", "Mobile No", contractor_info["Mobile_No"]])
        sheet.append(["Block", contractor_info["Block_Name"], " ", "Email", contractor_info["Email"]])
        sheet.append([])

        # Table Headers - include all hold types as separate columns
        base_headers = ["PMC No", "Village", "Work Type", "Invoice Details", "Invoice Date", "Invoice No",
                        "Basic Amount", "Debit", "After Debit Amount", "GST (18%)", "Amount", "TDS (1%)",
                        "SD (5%)", "On Commission", "Hydro Testing", "GST SD Amount"]

        hold_headers = [ht['hold_type'] for ht in hold_types]

        payment_headers = ["Final Amount", "Payment Amount", "TDS Payment", "Total Paid", "UTR"]

        sheet.append(base_headers + hold_headers + payment_headers)

        seen_invoices = set()
        seen_gst_notes = set()
        processed_payments = set()

        # Process invoices
        for inv in invoices:
            invoice_no = inv["Invoice_No"]
            payments = payments_map.get(invoice_no, [])

            # Process invoice row with first payment (if exists)
            if invoice_no not in seen_invoices:
                seen_invoices.add(invoice_no)
                first_payment = payments[0] if len(payments) > 0 else None

                # Base invoice data
                row = [
                    pmc_no, inv["Village_Name"], inv["Work_Type"], inv["Invoice_Details"],
                    inv["Invoice_Date"], invoice_no, inv["Basic_Amount"], inv["Debit_Amount"],
                    inv["After_Debit_Amount"], inv["GST_Amount"], inv["Amount"], inv["TDS_Amount"],
                    inv["SD_Amount"], inv["On_Commission"], inv["Hydro_Testing"], inv["GST_SD_Amount"]
                ]

                # Add hold amounts for each hold type
                invoice_holds = hold_data.get(inv["Invoice_Id"], {})
                for ht_id in hold_type_map.keys():
                    row.append(invoice_holds.get(ht_id, ""))

                # Add payment information
                row += [
                    inv["Final_Amount"],
                    first_payment["Payment_Amount"] if first_payment else "",
                    first_payment["TDS_Payment_Amount"] if first_payment else "",
                    first_payment["Total_amount"] if first_payment else "",
                    first_payment["UTR"] if first_payment else ""
                ]

                sheet.append(row)

                if first_payment:
                    payment_id = f"{invoice_no}-{first_payment['Payment_Amount']}-{first_payment.get('UTR', '')}"
                    processed_payments.add(payment_id)

            # Process GST release if exists (only if we have a matching GST record)
            if inv["gst_pmc_no"] and inv["gst_invoice_no"] and inv["gst_invoice_no"] not in seen_gst_notes:
                seen_gst_notes.add(inv["gst_invoice_no"])

                # Find the payment that matches this GST release
                gst_payment = None
                for payment in payments[1:]:  # Skip first payment (already used for invoice)
                    if payment['invoice_no'] == inv["gst_invoice_no"]:
                        gst_payment = payment
                        break

                # If no payment found in the invoice's payments, check all payments
                if not gst_payment:
                    gst_payments = payments_map.get(inv["gst_invoice_no"], [])
                    if gst_payments:
                        gst_payment = gst_payments[0]

                # GST release row
                row = [
                    pmc_no, "", "", "GST Release Note", "", inv["gst_invoice_no"],
                    inv["gst_basic_amount"], "", "", "", "", "", "", "", "", ""  # Empty GST SD Amount
                ]

                # Empty holds for GST release
                row += ["" for _ in hold_headers]

                # Add payment information
                row += [
                    inv["gst_final_amount"],
                    gst_payment["Payment_Amount"] if gst_payment else "",
                    gst_payment["TDS_Payment_Amount"] if gst_payment else "",
                    gst_payment["Total_amount"] if gst_payment else "",
                    gst_payment["UTR"] if gst_payment else ""
                ]

                sheet.append(row)

                if gst_payment:
                    payment_id = f"{inv['gst_invoice_no']}-{gst_payment['Payment_Amount']}-{gst_payment.get('UTR', '')}"
                    processed_payments.add(payment_id)

            # Process remaining payments as extra payments
            for payment in payments[1:]:
                payment_id = f"{payment['invoice_no']}-{payment['Payment_Amount']}-{payment.get('UTR', '')}"
                if payment_id not in processed_payments:
                    row = [
                        pmc_no, "", "", "", "", payment['invoice_no'],
                        "", "", "", "", "", "", "", "", "", ""  # Empty GST SD Amount
                    ]

                    # Empty holds for extra payments
                    row += ["" for _ in hold_headers]

                    # Add payment information
                    row += [
                        "",
                        payment["Payment_Amount"],
                        payment["TDS_Payment_Amount"],
                        payment["Total_amount"],
                        payment["UTR"]
                    ]

                    sheet.append(row)
                    processed_payments.add(payment_id)

        # Process extra payments (null invoice_no)
        for payment in extra_payments:
            payment_id = f"null-{payment['Payment_Amount']}-{payment.get('UTR', '')}"
            if payment_id not in processed_payments:
                row = [
                    pmc_no, "", "", "", "", "",
                    "", "", "", "", "", "", "", "", "", ""  # Empty GST SD Amount
                ]

                # Empty holds for null invoice payments
                row += ["" for _ in hold_headers]

                # Add payment information
                row += [
                    "",
                    payment["Payment_Amount"],
                    payment["TDS_Payment_Amount"],
                    payment["Total_amount"],
                    payment["UTR"]
                ]

                sheet.append(row)
                processed_payments.add(payment_id)

        # Calculate totals
        total_basic_amount = 0
        total_tds_amount = 0
        total_sd_amount = 0
        total_on_commission = 0
        total_hold_amount = 0
        total_final_amount = 0
        total_payment_amount = 0
        total_tds_payment_amount = 0
        total_total_paid = 0

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            try:
                total_basic_amount += float(row[6] or 0)  # Basic_Amount
                total_tds_amount += float(row[11] or 0)  # TDS_Amount
                total_sd_amount += float(row[12] or 0)  # SD_Amount
                total_on_commission += float(row[13] or 0)  # On_Commission
                total_final_amount += float(row[-5] or 0)  # Final_Amount
                total_payment_amount += float(row[-4] or 0)  # Payment_Amount
                total_tds_payment_amount += float(row[-3] or 0)  # TDS_Payment
                total_total_paid += float(row[-2] or 0)  # Total_Paid

                # Sum of hold amounts
                hold_start_col = len(base_headers)
                hold_end_col = hold_start_col + len(hold_headers)
                total_hold_amount += sum(float(row[i] or 0) for i in range(hold_start_col, hold_end_col))
            except (ValueError, IndexError, TypeError):
                continue

        # Append totals row
        totals_row = [
            "TOTAL", "", "", "", "", "",
            total_basic_amount, "", "", "", "", total_tds_amount, total_sd_amount,
            total_on_commission, "", "",  # Empty GST SD Amount
        ]

        # Add hold totals
        totals_row += [total_hold_amount] + [""] * (len(hold_headers) - 1)

        # Add payment totals
        totals_row += [
            total_final_amount,
            total_payment_amount,
            total_tds_payment_amount,
            total_total_paid,
            ""  # UTR column remains empty
        ]

        sheet.append([])
        sheet.append(totals_row)

        # Make totals row bold
        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)

        # Save Excel file
        workbook.save(output_file)
        workbook.close()

    finally:
        cursor.close()
        connection.close()

    return send_from_directory(output_folder, f"PMC_Report_{pmc_no}.xlsx", as_attachment=True)


# --------- Hold Types Controller --------------------------------------------
#  Route to Add a New Hold Type
@app.route('/add_hold_type', methods=['POST', 'GET'])
def add_hold_type():
    connection = config.get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch all hold types using the stored procedure
        cursor.callproc("GetAllHoldTypes")
        hold_types = []
        for hold in cursor.stored_results():
            hold_types = hold.fetchall()

        if request.method == 'POST':
            hold_type = request.form.get('hold_type', '').strip()

            # Validation: Must start with a letter
            if not hold_type or not hold_type[0].isalpha():
                return jsonify({"status": "error", "message": "Hold Type must start with a letter."}), 400

            # Validation: Check if it already exists (case-insensitive)
            cursor.execute("SELECT COUNT(*) AS count FROM hold_types WHERE LOWER(hold_type) = LOWER(%s)", (hold_type,))
            if cursor.fetchone()['count'] > 0:
                return jsonify({"status": "error", "message": "This Hold Type already exists."}), 400

            try:
                # Insert new hold type into the database
                # cursor.execute("INSERT INTO hold_types (hold_type) VALUES (%s)", (hold_type,))
                # connection.commit()
                cursor.callproc('SaveHoldType', [hold_type])
                connection.commit()

                return jsonify({"status": "success", "message": "Hold Type added successfully!"}), 201
            except mysql.connector.Error as e:
                connection.rollback()
                return jsonify({"status": "error", "message": f"Database error: {str(e)}"}), 500

    except mysql.connector.Error as e:
        return jsonify({"status": "error", "message": f"Database error: {str(e)}"}), 500

    finally:
        cursor.close()
        connection.close()

    return render_template('add_hold_type.html', Hold_Types_data=hold_types)

# Route to Update Hold Type
# @app.route('/update_hold_type/<int:id>', methods=['POST', 'GET'])
# def update_hold_type(id):
#     # GET request: Show the form with the current hold type
#     if request.method == 'GET':
#         connection = config.get_db_connection()
#         cursor = connection.cursor()
#         # cursor.execute("SELECT * FROM hold_types WHERE hold_type_id = %s", (id,))
#         # hold_type = cursor.fetchone()
#
#         cursor.callproc("GetHoldTypesById", (id,))
#         for hold in cursor.stored_results():
#             hold_type = hold.fetchone()
#
#         cursor.close()
#         connection.close()
#
#         if not hold_type:
#             return jsonify({'status': 'error', 'message': 'Hold Type not found.'}), 404
#
#         return render_template('edit_hold_type.html', hold_type=hold_type)
#
#     # POST request: Update the hold type
#     if request.method == 'POST':
#         new_hold_type = request.form.get('hold_type').strip()
#
#         # Validation: Must start with a letter
#         if not new_hold_type or not new_hold_type[0].isalpha():
#             return jsonify(ResponseHandler.invalid_name('Hold Type')), 400
#
#         connection = config.get_db_connection()
#         cursor = connection.cursor()
#
#         try:
#             # Check if the hold type exists before updating
#             # cursor.execute("SELECT * FROM hold_types WHERE hold_type_id = %s", (id,))
#             # hold_type = cursor.fetchone()
#             cursor.callproc("GetHoldTypesById", (id,))
#             for hold in cursor.stored_results():
#                 hold_type = hold.fetchone()
#
#             if not hold_type:
#                 return jsonify({'status': 'error', 'message': 'Hold Type not found.'}), 404
#
#             # Update the hold type
#             # cursor.execute("UPDATE hold_types SET hold_type = %s WHERE hold_type_id = %s", (new_hold_type, id))
#             cursor.callproc("UpdateHoldTypeById", (id,new_hold_type))
#             connection.commit()
#             return jsonify(ResponseHandler.update_success('Hold Type'))
#
#         except mysql.connector.Error as e:
#             connection.rollback()
#             return jsonify(ResponseHandler.update_failure('Hold Type')), 500
#         finally:
#             cursor.close()
#             connection.close()


@app.route('/update_hold_type/<int:id>', methods=['GET', 'POST'])
def update_hold_type(id):
    connection = config.get_db_connection()
    cursor = connection.cursor()

    if request.method == 'GET':
        cursor.callproc("GetHoldTypesById", (id,))
        for hold in cursor.stored_results():
            hold_type = hold.fetchone()
        cursor.close()
        connection.close()

        if not hold_type:
            return jsonify({'status': 'error', 'message': 'Hold Type not found.'}), 404

        return render_template('edit_hold_type.html', hold_type=hold_type)

    elif request.method == 'POST':
        new_hold_type = request.form.get('hold_type', '').strip()

        if not new_hold_type or not new_hold_type[0].isalpha():
            return jsonify(ResponseHandler.invalid_name('Hold Type')), 400

        try:
            cursor.callproc("GetHoldTypesById", (id,))
            for h in cursor.stored_results():
                hold_type = h.fetchone()

            if not hold_type:
                return jsonify({'status': 'error', 'message': 'Hold Type not found.'}), 404

            cursor.callproc("UpdateHoldTypeById", (id, new_hold_type))
            connection.commit()  # 🔥 Needed to actually save changes

            return jsonify(ResponseHandler.update_success('Hold Type')), 200

        except mysql.connector.Error as e:
            connection.rollback()
            return jsonify(ResponseHandler.update_failure('Hold Type')), 500

        finally:
            cursor.close()
            connection.close()


#  Route to Delete Hold Type
@app.route('/delete_hold_type/<int:id>', methods=['POST'])
def delete_hold_type(id):
    connection = config.get_db_connection()
    cursor = connection.cursor()

    try:
        # cursor.execute("SELECT * FROM hold_types WHERE hold_type_id = %s", (id,))
        # hold_type = cursor.fetchone()
        cursor.callproc("GetHoldTypesById", (id,))
        for hold in cursor.stored_results():
            hold_type = hold.fetchone()

        if not hold_type:
            return jsonify({'status': 'error', 'message': 'Hold Type not found.'}), 404

        # Proceed with deletion
        # cursor.execute("DELETE FROM hold_types WHERE hold_type_id = %s", (id,))
        cursor.callproc("DeleteHoldType", (id,))
        connection.commit()
        return jsonify(ResponseHandler.delete_success('Hold Type'))

    except mysql.connector.Error as e:
        return jsonify(ResponseHandler.delete_failure('Hold Type')), 500
    finally:
        cursor.close()
        connection.close()
# -- end hold types controlller --------------------


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)

